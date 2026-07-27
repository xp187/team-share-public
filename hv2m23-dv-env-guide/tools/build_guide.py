#!/usr/bin/env python3
"""Build the HV2M23 DV environment guide from the current source tree."""

from __future__ import annotations

import html
import json
import re
from fnmatch import fnmatchcase
from collections import Counter
from pathlib import Path
from textwrap import dedent
from urllib.parse import quote

from openpyxl import load_workbook


GUIDE_DIR = Path(__file__).resolve().parents[1]
ASSET_DIR = GUIDE_DIR / "assets"
TESTS_DIR = Path(r"E:\DV_TCON_C\top\tests")
SOURCE_ROOT = Path(r"E:\DV_TCON_C")
CASELIST_XLSX = Path(
    r"C:\Users\xiapeng2\Desktop\HV2M23\04.Architecture\IP_digital\2.Verification"
    r"\1.EDA\3.HV2M23_EDA_case_list.xlsx")
BUILD_DATE = "2026-07-27"

NAV_ITEMS = [
    ("index.html", "首页"),
    ("overview.html", "概览"),
    ("tb-arch.html", "TB 架构"),
    ("stimulus.html", "激励与 Golden"),
    ("checkers.html", "检查机制"),
    ("plan.html", "验证计划"),
    ("run.html", "运行与回归"),
    ("cases.html", "Case 计划索引"),
    ("portability.html", "复用与移植"),
    ("faq.html", "FAQ"),
]

CHECK_LABELS = {
    "data_merge_check_on": "Data Merge",
    "digital_top_check_on": "Digital Top",
    "chopper_check_on": "Chopper",
    "analog_check_on": "Analog",
    "drd_input_check_on": "DRD input",
    "drd_output_check_on": "DRD output",
    "i2c_reg_check_on": "I2C register",
    "digital_dplc_check": "DPLC",
}


def esc(value: object) -> str:
    """Escape text for HTML output."""
    return html.escape(str(value), quote=True)


def read_text(path: Path) -> str:
    """Read source text while tolerating legacy encodings."""
    for encoding in ("utf-8", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return ""


def parse_macros(text: str) -> dict[str, str]:
    """Parse active SystemVerilog preprocessor definitions."""
    macros: dict[str, str] = {}
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped.startswith("`define "):
            continue
        match = re.match(r"`define\s+(\w+)(?:\s+(.*?))?\s*$", stripped)
        if match:
            macros[match.group(1)] = (match.group(2) or "1").strip()
    return macros


def parse_assignments(text: str) -> dict[str, str]:
    """Parse direct rx_cfg/env_cfg assignments from a testcase."""
    assignments: dict[str, str] = {}
    pattern = re.compile(r"(?:rx_cfg|env_cfg)\.(\w+)\s*=\s*([^;\n]+)")
    for key, value in pattern.findall(text):
        assignments[key] = " ".join(value.split())
    return assignments


def classify_case(name: str, macros: dict[str, str], source: str) -> str:
    """Classify a testcase from its directory name and source features."""
    key = name.lower()
    if "drdod" in key or "drd_panel" in key or "DRDOD" in macros:
        return "DRDOD"
    if "dplc" in key or "DPLC" in macros:
        return "DPLC"
    if "i2c" in key or "isprx_access_reg" in key:
        return "Register / I2C"
    if any(token in key for token in ("wake", "pon_opt", "xon")):
        return "Power / Wake"
    if any(token in key for token in ("prefix", "training", "without_eol", "error_state", "bac_bac")):
        return "Protocol abnormal"
    if any(token in key for token in ("dbc", "chop", "vgma", "vbk", "tpd", "tpw")):
        return "Analog control"
    if any(token in key for token in ("unlock", "bwda", "bwdb", "bwdl", "utc")):
        return "Unlock / UTC"
    if re.search(r"t_H\d+_V\d+_F\d+_P\d+_L\d+", name):
        return "Video matrix"
    if any(token in key for token in ("align", "packpos", "clock1x2", "descram")):
        return "Link / Align"
    if any(token in key for token in ("litest", "debug", "testo", "hiz")):
        return "Test IO"
    if "force " in source:
        return "Directed waveform"
    return "Datapath"


def describe_case(
        name: str, category: str, macros: dict[str, str],
        assignments: dict[str, str], cfg_count: int) -> str:
    """Create a source-derived testcase description."""
    match = re.search(
        r"H(\d+)_V(\d+)_F(\d+)_P(\d+)_L(\d+)_CHSEL(\d+)_"
        r"POLC(\d+)_DOTC(\d+)_SHL(\d+)",
        name,
    )
    parts: list[str] = []
    if match:
        hact, vact, fps, ports, lanes, chsel, polc, dotc, shl = match.groups()
        parts.append(
            f"{hact}x{vact} @ {fps} Hz，{ports} ports，"
            f"PAIR_NUM={lanes}；寄存器映射 CHSEL={chsel}、POLC={polc}、"
            f"DOTC={dotc}、SHL={shl}。"
        )
    elif category == "DRDOD":
        parts.append("验证 DRDOD pattern 生成、门控映射、bypass 或跨帧状态行为。")
    elif category == "Register / I2C":
        parts.append("通过 I2C transaction 执行寄存器读写、默认值或 unlock/reset 检查。")
    elif category == "Power / Wake":
        parts.append("定向触发 power/wake 条件，并检查寄存器保持、默认值或恢复过程。")
    elif category == "Protocol abnormal":
        parts.append("在 pixel/setting 链路注入异常符号或时序，检查错误数据隔离与恢复。")
    elif category == "Analog control":
        parts.append("配置模拟控制寄存器并检查输出数据或相关控制波形。")
    elif category == "Unlock / UTC":
        parts.append("验证 unlock、UTC 或 bandwidth detection 相关状态与 lane 行为。")
    else:
        parts.append("验证基础 datapath、链路控制或定向波形场景。")

    compile_chip = macros.get("CHIP_SEL")
    if compile_chip is not None:
        parts.append(f"编译宏 CHIP_SEL={compile_chip}。")
    frame_num = macros.get("FRAME_NUM")
    if frame_num:
        parts.append(f"FRAME_NUM={frame_num}。")
    if cfg_count:
        parts.append(f"目录含 {cfg_count} 个 cfg_frame 文件。")
    if assignments.get("inject_traning_code") == "1":
        parts.append("源码启用 inject_traning_code。")
    if assignments.get("inject_bac_pol") == "1":
        parts.append("源码启用 inject_bac_pol。")
    if assignments.get("setting_without_eol") == "1":
        parts.append("源码启用 setting_without_eol。")
    return " ".join(parts)


def scan_source_cases() -> list[dict[str, object]]:
    """Scan all testcase directories and return browser-ready metadata."""
    active_names = {
        line.strip()
        for line in read_text(TESTS_DIR / "case_list.txt").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    }
    cases: list[dict[str, object]] = []
    for case_dir in sorted(TESTS_DIR.iterdir(), key=lambda path: path.name.lower()):
        if not case_dir.is_dir() or not case_dir.name.startswith("t_"):
            continue
        user_def_path = case_dir / "user_def.sv"
        macros = parse_macros(read_text(user_def_path)) if user_def_path.exists() else {}
        sv_files = [
            path for path in case_dir.glob("*.sv")
            if path.name not in {"user_def.sv", "test_lib.sv", "waves_dumper.sv"}
        ]
        preferred = case_dir / f"{case_dir.name}.sv"
        source_path = preferred if preferred.exists() else (
            max(sv_files, key=lambda path: path.stat().st_size) if sv_files else None
        )
        source = read_text(source_path) if source_path else ""
        assignments = parse_assignments(source)
        category = classify_case(case_dir.name, macros, source)
        cfg_files = sorted(case_dir.glob("cfg_frame*.txt"))
        pattern_files = sorted(case_dir.glob("pattern/*.ppm"))
        enabled_checks = [
            CHECK_LABELS[key]
            for key, value in assignments.items()
            if key in CHECK_LABELS and value in {"1", "1'b1"}
        ]
        disabled_checks = [
            CHECK_LABELS[key]
            for key, value in assignments.items()
            if key in CHECK_LABELS and value in {"0", "1'b0"}
        ]
        force_lines = []
        for line_number, line in enumerate(source.splitlines(), start=1):
            stripped = " ".join(line.strip().split())
            if stripped.startswith("force "):
                force_lines.append(f"L{line_number}: {stripped[:150]}")
            if len(force_lines) == 3:
                break
        check_calls = len(re.findall(r"\b(?:check_|compare_)\w*\s*\(", source))
        error_calls = len(re.findall(r"`uvm_(?:error|fatal)\b", source))
        feature_macros = [
            name for name in macros
            if name in {
                "DRDOD", "DRDOD_CYCLIC", "DPLC", "I2C_SIM",
                "RX_PHY_BEH", "TUBE_NO_DELAY", "PAGE_TEST",
                "ONE_PAIR", "TWO_PAIR", "THREE_PAIR", "FOUR_PAIR",
            }
        ]
        source_rel = (
            f"tests/{case_dir.name}/{source_path.name}" if source_path else "No testcase .sv"
        )
        cases.append({
            "name": case_dir.name,
            "category": category,
            "active": case_dir.name in active_names,
            "description": describe_case(
                case_dir.name, category, macros, assignments, len(cfg_files)),
            "source": source_rel,
            "macros": {
                key: macros[key]
                for key in (
                    "CHIP_SEL", "COLOR_DEPTH", "PORT_NUM", "PAIR_NUM",
                    "HACT", "VACT", "REF_RATE", "FRAME_NUM",
                )
                if key in macros
            },
            "features": feature_macros,
            "enabledChecks": enabled_checks,
            "disabledChecks": disabled_checks,
            "cfgCount": len(cfg_files),
            "patternCount": len(pattern_files),
            "forces": force_lines,
            "checkCalls": check_calls,
            "errorCalls": error_calls,
            "hasReadme": (case_dir / "README.md").exists(),
        })
    return cases


def clean_cell(value: object) -> str:
    """Normalize spreadsheet text without discarding line-level meaning."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return f"{value:.6f}".rstrip("0").rstrip(".")
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in text.splitlines()]
    return "\n".join(line for line in lines if line).strip()


def normalize_case_name(raw_name: str) -> tuple[str, list[str]]:
    """Extract the primary testcase name and aliases from an Excel cell."""
    candidates = re.findall(r"t_[A-Za-z0-9_]+", raw_name)
    if not candidates:
        return raw_name.splitlines()[0].strip(), []
    primary = candidates[0]
    aliases = list(dict.fromkeys(candidates[1:]))
    return primary, aliases


def polished_excel_text(text: str, fallback: str) -> str:
    """Turn spreadsheet notes into readable prose while preserving content."""
    if not text:
        return fallback
    return text.replace("\n", "；")


def link_register_cases(
        registers: list[dict[str, object]],
        plan: list[dict[str, object]]) -> None:
    """Link register Sim_Check entries to testcase names with traceable rules."""
    case_names = sorted({str(item["name"]) for item in plan})
    keyword_rules = {
        "SHL CASE": "SHL", "CHSEL CASE": "CHSEL",
        "DOTC CASE": "DOTC", "POLC CASE": "POLC",
        "H120V CASE": "H120V", "DRDOD CASE": "DRD",
        "SD_CHOP CASE": "SD_CHOP", "G_CHOP CASE": "G_CHOP",
        "DPLC CASE": "DPLC", "UTC CASE": "utc",
    }
    case_to_registers: dict[str, list[dict[str, str]]] = {
        name: [] for name in case_names}
    for register in registers:
        sim_check = str(register["simCheck"])
        matches: set[str] = set()
        method = ""
        patterns = re.findall(r"t_[A-Za-z0-9_*]+", sim_check)
        if patterns:
            method = "Excel testcase name/pattern"
            for pattern in patterns:
                matches.update(name for name in case_names
                               if fnmatchcase(name, pattern))
        elif sim_check in keyword_rules:
            method = f"Feature keyword: {keyword_rules[sim_check]}"
            keyword = keyword_rules[sim_check].lower()
            matches.update(name for name in case_names
                           if keyword in name.lower())
        register["caseNames"] = sorted(matches)
        register["linkMethod"] = method
        for name in matches:
            case_to_registers[name].append({
                "address": str(register["address"]),
                "name": str(register["name"]),
                "simCheck": sim_check,
                "linkMethod": method,
            })
    for item in plan:
        item["registers"] = case_to_registers[str(item["name"])]


def load_excel_plan(
        source_cases: list[dict[str, object]]) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Load canonical case-plan rows and supporting workbook sheets."""
    if not CASELIST_XLSX.exists():
        raise FileNotFoundError(f"Case list workbook not found: {CASELIST_XLSX}")
    workbook = load_workbook(CASELIST_XLSX, data_only=True)
    source_by_name = {str(item["name"]): item for item in source_cases}
    plan: list[dict[str, object]] = []

    for sheet_name, case_type in (("normal case", "Normal"),
                                  ("waveform case", "Waveform")):
        sheet = workbook[sheet_name]
        inherited = {"feature1": "", "feature2": "", "description": "",
                     "checkpoint": "", "comment": ""}
        for row_number in range(2, sheet.max_row + 1):
            raw_name = clean_cell(sheet.cell(row_number, 4).value)
            if not raw_name:
                continue
            values = {
                "feature1": clean_cell(sheet.cell(row_number, 1).value),
                "feature2": clean_cell(sheet.cell(row_number, 2).value),
                "description": clean_cell(sheet.cell(row_number, 5).value),
                "checkpoint": clean_cell(sheet.cell(row_number, 6).value),
                "comment": clean_cell(sheet.cell(row_number, 7).value),
            }
            if values["feature1"] or values["feature2"]:
                for key in ("description", "checkpoint", "comment"):
                    if not values[key]:
                        inherited[key] = ""
            for key, value in values.items():
                if value:
                    inherited[key] = value
            name, aliases = normalize_case_name(raw_name)
            source = source_by_name.get(name)
            if source is None:
                source = next((source_by_name.get(alias) for alias in aliases
                               if alias in source_by_name), None)
            category = inherited["feature1"] or inherited["feature2"] or case_type
            description = polished_excel_text(
                inherited["description"],
                f"验证 {category} 分类下该 {case_type} testcase 的目标功能与边界条件。")
            checkpoint = polished_excel_text(
                inherited["checkpoint"],
                "确认配置和激励进入目标 DUT 路径；检查关键状态、输出数据和恢复行为，并产生确定性的自动 PASS/FAIL 结果。")
            item: dict[str, object] = {
                "name": name,
                "rawName": raw_name,
                "aliases": aliases,
                "planId": f"{sheet_name}:{row_number}",
                "sheet": sheet_name,
                "row": row_number,
                "caseType": case_type,
                "category": category,
                "feature1": inherited["feature1"],
                "feature2": inherited["feature2"],
                "laneNote": clean_cell(sheet.cell(row_number, 3).value),
                "description": description,
                "checkpoint": checkpoint,
                "comment": polished_excel_text(inherited["comment"], ""),
                "runSummary": clean_cell(sheet.cell(row_number, 8).value),
                "owner": clean_cell(sheet.cell(row_number, 9).value),
                "revision": clean_cell(sheet.cell(row_number, 10).value),
                "status": clean_cell(sheet.cell(row_number, 11).value),
                "date": clean_cell(sheet.cell(row_number, 12).value),
                "linkedSource": source is not None,
            }
            if source:
                item.update({
                    "active": source["active"],
                    "source": source["source"],
                    "sourceDescription": source["description"],
                    "macros": source["macros"],
                    "features": source["features"],
                    "enabledChecks": source["enabledChecks"],
                    "disabledChecks": source["disabledChecks"],
                    "cfgCount": source["cfgCount"],
                    "patternCount": source["patternCount"],
                    "forces": source["forces"],
                    "checkCalls": source["checkCalls"],
                    "errorCalls": source["errorCalls"],
                })
            else:
                item.update({
                    "active": False, "source": "No matching testcase directory",
                    "sourceDescription": "", "macros": {}, "features": [],
                    "enabledChecks": [], "disabledChecks": [], "cfgCount": 0,
                    "patternCount": 0, "forces": [], "checkCalls": 0,
                    "errorCalls": 0,
                })
            plan.append(item)

    registers = []
    for row in workbook["register"].iter_rows(min_row=4, values_only=True):
        if row[2] is None:
            continue
        registers.append({
            "address": clean_cell(row[1]), "name": clean_cell(row[2]),
            "type": clean_cell(row[3]), "default": clean_cell(row[4]),
            "block": clean_cell(row[5]), "description": clean_cell(row[6]),
            "values": clean_cell(row[7]), "access": clean_cell(row[8]),
            "simCheck": clean_cell(row[9]),
        })
    link_register_cases(registers, plan)
    changes = []
    for row in workbook["ENV changelist"].iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        changes.append({
            "date": clean_cell(row[0]), "file": clean_cell(row[1]),
            "action": clean_cell(row[2]), "owner": clean_cell(row[3]),
            "status": clean_cell(row[4]), "note": clean_cell(row[5]),
        })
    formats = []
    video_sheet = workbook["video format"]
    for row_number in range(6, video_sheet.max_row + 1):
        row = [video_sheet.cell(row_number, column).value for column in range(1, 16)]
        if row[1] is None:
            continue
        formats.append({
            "no": clean_cell(row[1]), "hPeriod": clean_cell(row[2]),
            "vPeriod": clean_cell(row[3]), "hBlank": clean_cell(row[4]),
            "vBlank": clean_cell(row[5]), "fps": clean_cell(row[6]),
            "pixelClock": clean_cell(row[7]), "depth": clean_cell(row[8]),
            "bandwidth": clean_cell(row[9]), "channel": clean_cell(row[10]),
            "pcs": clean_cell(row[11]), "pair": clean_cell(row[12]),
            "ispSpeed": clean_cell(row[13]), "mode": clean_cell(row[14]),
        })
    coverage_row = list(workbook["coverage"].iter_rows(min_row=2, max_row=2,
                                                        values_only=True))[0]
    history = []
    for row in workbook["History"].iter_rows(min_row=3, values_only=True):
        if not any(value is not None for value in row):
            continue
        history.append({
            "date": clean_cell(row[1]), "comment": clean_cell(row[2]),
            "owner": clean_cell(row[3]),
        })
    case_status = []
    status_sheet = workbook["case status "]
    for row in status_sheet.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        case_status.append({
            "type": clean_cell(row[0]), "pass": clean_cell(row[1]),
            "error": clean_cell(row[2]), "ongoing": clean_cell(row[3]),
            "unbuilt": clean_cell(row[4]), "total": clean_cell(row[5]),
        })
    feature_groups = []
    for sheet_name, case_type in (("normal case", "Normal"),
                                  ("waveform case", "Waveform")):
        sheet = workbook[sheet_name]
        current_feature1 = ""
        for row_number in range(2, sheet.max_row + 1):
            feature1 = clean_cell(sheet.cell(row_number, 1).value)
            feature2 = clean_cell(sheet.cell(row_number, 2).value)
            if feature1:
                current_feature1 = feature1
            if not feature2:
                continue
            named_count = sum(
                1 for item in plan
                if item["sheet"] == sheet_name
                and item["feature1"] == current_feature1
                and item["feature2"] == feature2)
            feature_groups.append({
                "caseType": case_type, "feature1": current_feature1,
                "feature2": feature2, "namedCount": named_count,
                "source": f"{sheet_name}:{row_number}",
            })
    metadata = {
        "registers": registers,
        "changes": changes,
        "formats": formats,
        "history": history,
        "caseStatus": case_status,
        "featureGroups": feature_groups,
        "coverage": {
            "score": clean_cell(coverage_row[0]), "line": clean_cell(coverage_row[1]),
            "condition": clean_cell(coverage_row[2]), "toggle": clean_cell(coverage_row[3]),
            "fsm": clean_cell(coverage_row[4]), "branch": clean_cell(coverage_row[5]),
            "date": clean_cell(coverage_row[6]),
        },
        "workbook": str(CASELIST_XLSX),
    }
    return plan, metadata


CSS = r"""
:root {
  --bg: #f4f6f8;
  --surface: #ffffff;
  --surface-2: #eef2f5;
  --ink: #18212b;
  --muted: #607080;
  --line: #d6dde3;
  --blue: #1769aa;
  --blue-soft: #e4f1fa;
  --green: #167a55;
  --green-soft: #e2f3eb;
  --amber: #9a5b08;
  --amber-soft: #fff1d6;
  --red: #b83a3a;
  --code: #17212b;
  --code-ink: #e7edf2;
  --shadow: 0 8px 24px rgba(24, 33, 43, .08);
}
@media (prefers-color-scheme: dark) {
  :root {
    --bg: #10161d;
    --surface: #18212b;
    --surface-2: #222e39;
    --ink: #edf3f7;
    --muted: #9cabb8;
    --line: #344451;
    --blue: #65b7ee;
    --blue-soft: #17364c;
    --green: #68cda3;
    --green-soft: #163b2e;
    --amber: #efbd6c;
    --amber-soft: #483317;
    --red: #f07c7c;
    --code: #0c1117;
    --code-ink: #edf3f7;
    --shadow: 0 8px 28px rgba(0, 0, 0, .28);
  }
}
* { box-sizing: border-box; }
html { scroll-behavior: smooth; }
html, body { max-width: 100%; overflow-x: hidden; }
body {
  margin: 0;
  background: var(--bg);
  color: var(--ink);
  font: 15px/1.7 Inter, "Segoe UI", "Microsoft YaHei", sans-serif;
}
a { color: var(--blue); }
.topbar {
  position: sticky;
  top: 0;
  z-index: 20;
  border-bottom: 1px solid var(--line);
  background: color-mix(in srgb, var(--surface) 94%, transparent);
  backdrop-filter: blur(12px);
}
.topbar-inner {
  width: 100%;
  max-width: 1240px;
  margin: auto;
  padding: 10px 20px;
  display: flex;
  align-items: center;
  gap: 18px;
}
.brand { color: var(--ink); font-weight: 750; text-decoration: none; white-space: nowrap; }
.brand span { color: var(--blue); }
.nav { display: flex; gap: 2px; overflow-x: auto; scrollbar-width: none; }
.nav a { color: var(--muted); padding: 6px 9px; text-decoration: none; white-space: nowrap; border-radius: 5px; font-size: 13px; }
.nav a:hover, .nav a.active { background: var(--blue-soft); color: var(--blue); }
.layout { width: 100%; max-width: 1240px; margin: auto; padding: 30px 20px 64px; }
.content-grid { display: grid; grid-template-columns: minmax(0, 1fr); align-items: start; }
.page-content { min-width: 0; }
.section-nav { position: fixed; z-index: 30; left: 12px; top: 86px; width: 40px; max-height: 54px; overflow: hidden; border: 1px solid var(--line); border-left: 3px solid var(--blue); border-radius: 5px; padding: 10px 7px; background: var(--surface); box-shadow: var(--shadow); transition: width 160ms ease, max-height 160ms ease; }
.section-nav:hover, .section-nav:focus-within { width: 220px; max-height: calc(100vh - 106px); overflow-y: auto; }
.section-nav strong { display: block; width: 24px; margin: 0 auto 8px; font-size: 12px; color: var(--blue); line-height: 1.3; text-align: center; }
.section-nav:hover strong, .section-nav:focus-within strong { width: auto; text-align: left; margin: 0 7px 8px; }
.section-nav a { display: block; padding: 5px 7px; color: var(--muted); text-decoration: none; font-size: 12px; line-height: 1.35; border-radius: 4px; }
.section-nav:not(:hover):not(:focus-within) a { opacity: 0; pointer-events: none; }
.section-nav a:hover { color: var(--blue); background: var(--blue-soft); }
h2[id] { scroll-margin-top: 76px; }
.page-head { margin-bottom: 28px; border-bottom: 1px solid var(--line); padding-bottom: 22px; }
.eyebrow { color: var(--blue); font-size: 12px; font-weight: 750; text-transform: uppercase; }
h1 { margin: 6px 0 8px; font-size: clamp(28px, 4vw, 42px); line-height: 1.18; letter-spacing: 0; }
.lead { max-width: 820px; color: var(--muted); font-size: 16px; }
.source-stamp { margin-top: 12px; color: var(--muted); font-size: 12px; }
h2 { margin: 42px 0 16px; font-size: 23px; line-height: 1.25; }
h3 { margin: 26px 0 10px; font-size: 17px; }
p { margin: 8px 0 14px; }
.stats { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin: 20px 0 30px; }
.stat { border-left: 3px solid var(--blue); background: var(--surface); padding: 14px 16px; box-shadow: var(--shadow); }
.stat strong { display: block; font-size: 24px; line-height: 1.2; }
.stat span { color: var(--muted); font-size: 12px; }
.grid-2, .grid-3 { display: grid; gap: 14px; }
.grid-2 { grid-template-columns: repeat(2, minmax(0, 1fr)); }
.grid-3 { grid-template-columns: repeat(3, minmax(0, 1fr)); }
.panel, .case-card { background: var(--surface); border: 1px solid var(--line); border-radius: 6px; padding: 18px; }
.panel h3, .case-card h3 { margin-top: 0; }
.panel.accent { border-left: 4px solid var(--green); }
.note { border-left: 4px solid var(--amber); background: var(--amber-soft); padding: 13px 16px; margin: 18px 0; }
.source { display: inline-block; color: var(--muted); font-family: Consolas, monospace; font-size: 12px; }
.figure { margin: 18px 0 28px; border: 1px solid var(--line); background: var(--surface); padding: 14px; overflow-x: auto; }
.figure img { display: block; width: 100%; min-width: 720px; height: auto; }
.caption { color: var(--muted); font-size: 12px; text-align: center; padding-top: 8px; }
table { width: 100%; border-collapse: collapse; margin: 12px 0 22px; background: var(--surface); }
.table-scroll { width: 100%; overflow-x: auto; }
.table-scroll table { min-width: 900px; }
.register-table { min-width: 1800px !important; table-layout: fixed; }
.register-table th:nth-child(1), .register-table td:nth-child(1) { width: 90px; }
.register-table th:nth-child(2), .register-table td:nth-child(2) { width: 150px; }
.register-table th:nth-child(3), .register-table td:nth-child(3),
.register-table th:nth-child(4), .register-table td:nth-child(4),
.register-table th:nth-child(5), .register-table td:nth-child(5) { width: 105px; }
.register-table th:nth-child(6), .register-table td:nth-child(6) { width: 250px; }
.register-table th:nth-child(7), .register-table td:nth-child(7) { width: 300px; }
.register-table th:nth-child(8), .register-table td:nth-child(8) { width: 180px; }
.register-table th:nth-child(9), .register-table td:nth-child(9) { width: 260px; }
.register-table th:nth-child(10), .register-table td:nth-child(10) { width: 180px; }
.register-case-links { max-height: 220px; overflow-y: auto; }
.register-case-links a { display: block; margin: 4px 0; }
details > summary { cursor: pointer; color: var(--blue); font-weight: 650; margin: 10px 0; }
th, td { text-align: left; vertical-align: top; border-bottom: 1px solid var(--line); padding: 10px 12px; }
th { color: var(--muted); font-size: 12px; text-transform: uppercase; }
code { font: 12px/1.5 Consolas, monospace; background: var(--blue-soft); color: var(--blue); padding: 2px 5px; border-radius: 3px; word-break: break-word; }
pre { background: var(--code); color: var(--code-ink); padding: 16px; overflow-x: auto; border-radius: 5px; }
pre code { background: transparent; color: inherit; padding: 0; }
.steps { counter-reset: step; padding: 0; list-style: none; }
.steps li { position: relative; padding: 0 0 20px 42px; }
.steps li::before { counter-increment: step; content: counter(step); position: absolute; left: 0; top: 0; width: 27px; height: 27px; border-radius: 50%; background: var(--blue); color: white; text-align: center; line-height: 27px; font-weight: 700; }
.steps li:not(:last-child)::after { content: ""; position: absolute; left: 13px; top: 29px; bottom: 2px; border-left: 1px solid var(--line); }
.tags { display: flex; gap: 5px; flex-wrap: wrap; margin: 8px 0; }
.tag { display: inline-flex; align-items: center; border: 1px solid var(--line); color: var(--muted); padding: 2px 7px; border-radius: 999px; font-size: 11px; }
.tag.active { border-color: var(--green); background: var(--green-soft); color: var(--green); }
.tag.warn { border-color: var(--amber); background: var(--amber-soft); color: var(--amber); }
.case-controls { position: sticky; top: 56px; z-index: 10; display: grid; grid-template-columns: minmax(220px, 1fr) repeat(4, minmax(130px, 170px)); gap: 8px; padding: 12px; background: var(--surface); border: 1px solid var(--line); box-shadow: var(--shadow); }
input, select { width: 100%; min-height: 38px; border: 1px solid var(--line); background: var(--surface-2); color: var(--ink); padding: 7px 10px; border-radius: 4px; font: inherit; }
.case-results { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 10px; margin-top: 14px; }
.case-title { display: flex; align-items: flex-start; gap: 8px; justify-content: space-between; }
.case-title code { font-size: 12px; }
.case-description { color: var(--muted); font-size: 13px; }
.case-card h4 { margin: 13px 0 3px; font-size: 12px; color: var(--ink); }
.case-checkpoint { margin: 3px 0 10px; padding: 9px 11px; border-left: 3px solid var(--green); background: var(--green-soft); font-size: 13px; white-space: pre-line; }
.case-evidence { margin-top: 9px; font-size: 12px; }
.case-evidence summary { cursor: pointer; color: var(--blue); }
.empty { padding: 40px; text-align: center; color: var(--muted); border: 1px dashed var(--line); }
.footer { margin-top: 60px; padding-top: 20px; border-top: 1px solid var(--line); color: var(--muted); font-size: 12px; }
@media (max-width: 840px) {
  .topbar-inner { align-items: flex-start; flex-direction: column; gap: 6px; }
  .topbar-inner, .nav { min-width: 0; }
  .nav { align-self: stretch; max-width: 100%; }
  .layout { padding: 22px 14px 48px; }
  p, li, td, .lead, .note, .case-description { overflow-wrap: anywhere; word-break: break-word; }
  .stats, .grid-2, .grid-3, .case-results { grid-template-columns: 1fr; }
  .case-controls { top: 93px; grid-template-columns: 1fr; }
  .content-grid { grid-template-columns: 1fr; gap: 12px; }
  .section-nav { position: static; max-height: 150px; border: 1px solid var(--line); padding: 10px; background: var(--surface); }
  .section-nav { width: auto; max-height: 150px; overflow-y: auto; }
  .section-nav:hover, .section-nav:focus-within { width: auto; }
  .section-nav strong, .section-nav:hover strong { width: auto; text-align: left; margin: 0 7px 8px; }
  .section-nav:not(:hover):not(:focus-within) a { opacity: 1; pointer-events: auto; }
  .section-nav a { display: inline-block; margin: 2px; }
  h1 { font-size: 29px; }
  .figure { margin-left: -14px; margin-right: -14px; border-left: 0; border-right: 0; }
}
"""


GUIDE_JS = r"""
(() => {
  const search = document.querySelector('#case-search');
  if (!search || !window.HV2_CASES) return;
  const caseType = document.querySelector('#case-type');
  const feature1 = document.querySelector('#case-feature1');
  const feature2 = document.querySelector('#case-feature2');
  const scope = document.querySelector('#case-scope');
  const results = document.querySelector('#case-results');
  const count = document.querySelector('#case-count');
  const initialQuery = new URLSearchParams(window.location.search).get('q');
  if (initialQuery) search.value = initialQuery;

  const escapeHtml = value => String(value)
    .replaceAll('&', '&amp;').replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;').replaceAll('"', '&quot;');

  const tags = values => values.map(value =>
    `<span class="tag">${escapeHtml(value)}</span>`).join('');

  function render() {
    const query = search.value.trim().toLowerCase();
    const filtered = window.HV2_CASES.filter(item => {
      const haystack = JSON.stringify(item).toLowerCase();
      return (!query || haystack.includes(query))
        && (!caseType.value || item.caseType === caseType.value)
        && (!feature1.value || (feature1.value === '__blank__'
          ? !item.feature1 : item.feature1 === feature1.value))
        && (!feature2.value || item.feature2 === feature2.value)
        && (scope.value !== 'pass' || item.status.toUpperCase() === 'PASS')
        && (scope.value !== 'linked' || item.linkedSource)
        && (scope.value !== 'missing' || !item.linkedSource);
    });
    count.textContent = `${filtered.length} / ${window.HV2_CASES.length}`;
    if (!filtered.length) {
      results.innerHTML = '<div class="empty">没有匹配的 testcase</div>';
      return;
    }
    results.innerHTML = filtered.map(item => {
      const macroTags = Object.entries(item.macros)
        .map(([key, value]) => `${key}=${value}`);
      const evidence = [
        `Excel: ${item.planId}`,
        `Source: ${item.source}`,
        `cfg_frame files: ${item.cfgCount}`,
        `pattern files: ${item.patternCount}`,
        `check/compare calls: ${item.checkCalls}`,
        `uvm_error/fatal calls: ${item.errorCalls}`,
        ...item.forces,
      ];
      return `<article class="case-card">
        <div class="case-title">
          <code>${escapeHtml(item.name)}</code>
          <span class="tag ${item.status.toUpperCase() === 'PASS' ? 'active' : ''}">${escapeHtml(item.status || 'status not set')}</span>
        </div>
        <div class="tags"><span class="tag warn">Case Type: ${escapeHtml(item.caseType)}</span><span class="tag">Feature I: ${escapeHtml(item.feature1 || 'Excel 未填写')}</span><span class="tag">Feature II: ${escapeHtml(item.feature2 || 'Excel 未填写')}</span><span class="tag ${item.linkedSource ? 'active' : 'warn'}">${item.linkedSource ? 'source linked' : 'Excel only'}</span></div>
        <h4>验证目标</h4><p class="case-description">${escapeHtml(item.description)}</p>
        <h4>检查点</h4><p class="case-checkpoint">${escapeHtml(item.checkpoint)}</p>
        ${item.comment ? `<h4>说明</h4><p class="case-description">${escapeHtml(item.comment)}</p>` : ''}
        ${item.registers.length ? `<h4>关联寄存器</h4><div class="tags">${tags(item.registers.map(reg => `${reg.address} ${reg.name}`))}</div>` : ''}
        <div class="tags">${item.owner ? `<span class="tag">Owner: ${escapeHtml(item.owner)}</span>` : ''}${item.date ? `<span class="tag">Date: ${escapeHtml(item.date)}</span>` : ''}${item.runSummary ? `<span class="tag">Run: ${escapeHtml(item.runSummary)}</span>` : ''}</div>
        <div class="tags">${tags(item.features)}${tags(macroTags)}</div>
        <div class="tags">${tags(item.enabledChecks.map(v => `ON: ${v}`))}${tags(item.disabledChecks.map(v => `OFF: ${v}`))}</div>
        ${item.sourceDescription ? `<details class="case-evidence"><summary>查看源码补充描述</summary><p>${escapeHtml(item.sourceDescription)}</p></details>` : ''}
        <details class="case-evidence"><summary>查看 Excel 与源码证据</summary><pre><code>${escapeHtml(evidence.join('\n'))}</code></pre></details>
      </article>`;
    }).join('');
  }

  [search, caseType, feature1, feature2, scope].forEach(control =>
    control.addEventListener('input', render));
  render();
})();
"""


def svg_document(width: int, height: int, content: str, label: str) -> str:
    """Wrap diagram content in a standalone responsive SVG."""
    return dedent(f"""\
        <!-- Auto-generated by tools/build_guide.py -->
        <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="{label}">
          <style>
            text {{ font-family: 'Helvetica Neue', Helvetica, Arial, 'PingFang SC', 'Microsoft YaHei', sans-serif; }}
            .bg {{ fill: #ffffff; }}
            .box {{ fill: #ffffff; stroke: #d1d5db; stroke-width: 1.5; }}
            .blue {{ fill: #eff6ff; stroke: #bfdbfe; stroke-width: 1.5; }}
            .green {{ fill: #f0fdf4; stroke: #bbf7d0; stroke-width: 1.5; }}
            .amber {{ fill: #fff7ed; stroke: #fed7aa; stroke-width: 1.5; }}
            .purple {{ fill: #faf5ff; stroke: #e9d5ff; stroke-width: 1.5; }}
            .ink {{ fill: #111827; font-size: 14px; font-weight: 600; }}
            .small {{ fill: #6b7280; font: 12px Consolas, monospace; }}
            .arrow {{ stroke: #2563eb; stroke-width: 2; fill: none; marker-end: url(#arrow-blue); }}
            .arrow-green {{ stroke: #16a34a; stroke-width: 2; fill: none; marker-end: url(#arrow-green); }}
            .arrow-orange {{ stroke: #ea580c; stroke-width: 1.7; fill: none; marker-end: url(#arrow-orange); }}
            .arrow-purple {{ stroke: #9333ea; stroke-width: 1.7; fill: none; marker-end: url(#arrow-purple); }}
            .dash {{ stroke-dasharray: 5 4; }}
          </style>
          <defs>
            <marker id="arrow-blue" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="7" markerHeight="7" orient="auto"><path d="M0 0 L10 5 L0 10 Z" fill="#2563eb"/></marker>
            <marker id="arrow-green" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="7" markerHeight="7" orient="auto"><path d="M0 0 L10 5 L0 10 Z" fill="#16a34a"/></marker>
            <marker id="arrow-orange" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="7" markerHeight="7" orient="auto"><path d="M0 0 L10 5 L0 10 Z" fill="#ea580c"/></marker>
            <marker id="arrow-purple" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="7" markerHeight="7" orient="auto"><path d="M0 0 L10 5 L0 10 Z" fill="#9333ea"/></marker>
          </defs>
          <rect class="bg" width="{width}" height="{height}"/>
          {content}
        </svg>
    """)


def build_svgs() -> None:
    """Write source-aligned architecture and workflow diagrams."""
    architecture = """
      <text class="ink" x="36" y="35">1. Static TB / HDL domain</text>
      <rect class="blue" x="35" y="55" width="220" height="105" rx="8"/><text class="ink" x="145" y="83" text-anchor="middle">chip_tb_top</text><text class="small" x="145" y="106" text-anchor="middle">DUT + clocks/resets</text><text class="small" x="145" y="127" text-anchor="middle">isptx_if[0..3] / i2c_if[0..3]</text><text class="small" x="145" y="148" text-anchor="middle">checker tap interfaces</text>
      <rect class="green" x="330" y="55" width="220" height="105" rx="8"/><text class="ink" x="440" y="83" text-anchor="middle">HV2M23 DUT</text><text class="small" x="440" y="106" text-anchor="middle">source input links</text><text class="small" x="440" y="127" text-anchor="middle">data merge / digital top</text><text class="small" x="440" y="148" text-anchor="middle">chopper / analog / DRDOD taps</text>
      <rect class="amber" x="625" y="55" width="280" height="105" rx="8"/><text class="ink" x="765" y="83" text-anchor="middle">uvm_config_db interface binding</text><text class="small" x="765" y="106" text-anchor="middle">vif, vif0..vif3</text><text class="small" x="765" y="127" text-anchor="middle">ana_data_output_if</text><text class="small" x="765" y="148" text-anchor="middle">data_merge / digital_top / chopper intf</text>
      <path class="arrow" d="M255 107 H330"/><path class="arrow-orange dash" d="M145 160 V182 H765 V160"/><text class="small" x="465" y="176">publishes virtual interfaces</text>

      <text class="ink" x="36" y="210">2. Test and configuration domain</text>
      <rect class="purple" x="35" y="230" width="220" height="120" rx="8"/><text class="ink" x="145" y="258" text-anchor="middle">Concrete testcase</text><text class="small" x="145" y="281" text-anchor="middle">extends base_test</text><text class="small" x="145" y="302" text-anchor="middle">creates/configures rx_cfg</text><text class="small" x="145" y="323" text-anchor="middle">config_db: uvm_test_top.*</text><text class="small" x="145" y="344" text-anchor="middle">starts base_vseq</text>
      <rect class="blue" x="330" y="230" width="250" height="120" rx="8"/><text class="ink" x="455" y="258" text-anchor="middle">base_test</text><text class="small" x="455" y="281" text-anchor="middle">cov_comp + sd_vsqr</text><text class="small" x="455" y="302" text-anchor="middle">sd_env + sd_env_1..3</text><text class="small" x="455" y="323" text-anchor="middle">count selected by CONNECT_NUM</text><text class="small" x="455" y="344" text-anchor="middle">connects ISPTX/I2C sequencers</text>
      <rect class="box" x="655" y="230" width="250" height="120" rx="8"/><text class="ink" x="780" y="258" text-anchor="middle">sd_vsqr + base_vseq</text><text class="small" x="780" y="281" text-anchor="middle">p_isptx_sqr[0..3]</text><text class="small" x="780" y="302" text-anchor="middle">p_i2c_sqr[0..3]</text><text class="small" x="780" y="323" text-anchor="middle">fork ISPTX sequence[0..3]</text><text class="small" x="780" y="344" text-anchor="middle">selected by CONNECT_NUM</text>
      <path class="arrow-orange" d="M255 290 H330"/><text class="small" x="268" y="282">factory build</text><path class="arrow" d="M580 290 H655"/><text class="small" x="594" y="282">handles</text>

      <text class="ink" x="36" y="405">3. Per-connection UVM environment (replicated 1..4 times)</text>
      <rect class="box dash" x="35" y="425" width="870" height="205" rx="8"/>
      <rect class="blue" x="65" y="465" width="180" height="100" rx="8"/><text class="ink" x="155" y="494" text-anchor="middle">ISPTX agent</text><text class="small" x="155" y="517" text-anchor="middle">sequencer + driver</text><text class="small" x="155" y="538" text-anchor="middle">setting/pixel traffic</text><text class="small" x="155" y="559" text-anchor="middle">uses isptx_if vif</text>
      <rect class="amber" x="285" y="465" width="180" height="100" rx="8"/><text class="ink" x="375" y="494" text-anchor="middle">I2C agent</text><text class="small" x="375" y="517" text-anchor="middle">sequencer + driver</text><text class="small" x="375" y="538" text-anchor="middle">monitor + scoreboard</text><text class="small" x="375" y="559" text-anchor="middle">uses i2c_if vif</text>
      <rect class="purple" x="505" y="465" width="180" height="100" rx="8"/><text class="ink" x="595" y="494" text-anchor="middle">checker_agent</text><text class="small" x="595" y="517" text-anchor="middle">6 monitor families</text><text class="small" x="595" y="538" text-anchor="middle">6 scoreboards</text><text class="small" x="595" y="559" text-anchor="middle">analysis connections</text>
      <rect class="green" x="725" y="465" width="150" height="100" rx="8"/><text class="ink" x="800" y="494" text-anchor="middle">rx_cfg</text><text class="small" x="800" y="517" text-anchor="middle">dimensions / id</text><text class="small" x="800" y="538" text-anchor="middle">checker gates</text><text class="small" x="800" y="559" text-anchor="middle">frame settings</text>
      <path class="arrow-orange dash" d="M145 350 V440 H800 V465"/><path class="arrow" d="M655 320 H620 V450 H155 V465"/><path class="arrow" d="M655 335 H640 V445 H375 V465"/>
      <path class="arrow" d="M155 465 H18 V92 H35"/><path class="arrow" d="M375 465 H18"/>
      <path class="arrow-green" d="M480 160 V190 H615 V450 H595 V465"/>
      <path class="arrow-orange dash" d="M765 160 H930 V440 H155 V465"/><path class="arrow-orange dash" d="M375 440 V465"/><path class="arrow-orange dash" d="M595 440 V465"/>
      <path class="arrow-orange dash" d="M800 565 V605 H155 V565"/><path class="arrow-orange dash" d="M375 605 V565"/><path class="arrow-orange dash" d="M595 605 V565"/><text class="small" x="445" y="622">rx_cfg consumed by sequence, driver, monitors and scoreboards</text>

      <rect class="amber" x="950" y="230" width="215" height="120" rx="8"/><text class="ink" x="1058" y="258" text-anchor="middle">CONNECT_NUM</text><text class="small" x="1058" y="281" text-anchor="middle">env instance count</text><text class="small" x="1058" y="302" text-anchor="middle">sequence count</text><text class="small" x="1058" y="323" text-anchor="middle">interface index count</text><text class="small" x="1058" y="344" text-anchor="middle">actual control macro</text>
      <rect class="box" x="950" y="425" width="215" height="140" rx="8"/><text class="ink" x="1058" y="453" text-anchor="middle">PAIR_NUM / PORT_NUM</text><text class="small" x="1058" y="476" text-anchor="middle">frequency calculation</text><text class="small" x="1058" y="497" text-anchor="middle">pixel/data partition</text><text class="small" x="1058" y="518" text-anchor="middle">PPM width / model args</text><text class="small" x="1058" y="539" text-anchor="middle">not env instance count</text>
      <path class="arrow-orange dash" d="M950 290 H905"/><path class="arrow-purple dash" d="M950 495 H905"/>
      <g transform="translate(40 675)"><path class="arrow" d="M0 8 H28"/><text class="small" x="36" y="12">sequence/traffic</text><path class="arrow-green" d="M185 8 H213"/><text class="small" x="221" y="12">data/check path</text><path class="arrow-orange dash" d="M385 8 H413"/><text class="small" x="421" y="12">configuration/control</text><path class="arrow-purple dash" d="M640 8 H668"/><text class="small" x="676" y="12">data-layout parameter</text></g>
    """
    flow = """
      <text class="ink" x="35" y="35">A. Select and compile</text>
      <rect class="box" x="35" y="55" width="220" height="110" rx="8"/><text class="ink" x="145" y="83" text-anchor="middle">Testcase directory</text><text class="small" x="145" y="106" text-anchor="middle">test class + user_def.sv</text><text class="small" x="145" y="127" text-anchor="middle">cfg_frame*.txt + pattern</text><text class="small" x="145" y="148" text-anchor="middle">test_lib include selection</text>
      <rect class="blue" x="330" y="55" width="220" height="110" rx="8"/><text class="ink" x="440" y="83" text-anchor="middle">run_tc.sh</text><text class="small" x="440" y="106" text-anchor="middle">compile/elaborate TOP</text><text class="small" x="440" y="127" text-anchor="middle">chip_tb_top + selected test</text><text class="small" x="440" y="148" text-anchor="middle">SIM_TEST_PATH plusarg</text>
      <rect class="green" x="625" y="55" width="240" height="110" rx="8"/><text class="ink" x="745" y="83" text-anchor="middle">Elaborated simulation</text><text class="small" x="745" y="106" text-anchor="middle">DUT + static interfaces</text><text class="small" x="745" y="127" text-anchor="middle">CONNECT_NUM hierarchy</text><text class="small" x="745" y="148" text-anchor="middle">run_test(concrete testcase)</text>
      <path class="arrow-orange" d="M255 110 H330"/><path class="arrow-orange" d="M550 110 H625"/>

      <text class="ink" x="35" y="215">B. UVM build and connect</text>
      <rect class="purple" x="35" y="235" width="220" height="120" rx="8"/><text class="ink" x="145" y="263" text-anchor="middle">testcase.build_phase</text><text class="small" x="145" y="286" text-anchor="middle">configure rx_cfg</text><text class="small" x="145" y="307" text-anchor="middle">set uvm_test_top.*</text><text class="small" x="145" y="328" text-anchor="middle">checker gates / dimensions</text><text class="small" x="145" y="349" text-anchor="middle">id and protocol controls</text>
      <rect class="blue" x="330" y="235" width="220" height="120" rx="8"/><text class="ink" x="440" y="263" text-anchor="middle">base_test + env build</text><text class="small" x="440" y="286" text-anchor="middle">sd_env[0..CONNECT_NUM-1]</text><text class="small" x="440" y="307" text-anchor="middle">ISPTX + I2C + checker</text><text class="small" x="440" y="328" text-anchor="middle">sd_vsqr + coverage</text><text class="small" x="440" y="349" text-anchor="middle">all checker components created</text>
      <rect class="amber" x="625" y="235" width="240" height="120" rx="8"/><text class="ink" x="745" y="263" text-anchor="middle">connect/config binding</text><text class="small" x="745" y="286" text-anchor="middle">sequencer handles -> sd_vsqr</text><text class="small" x="745" y="307" text-anchor="middle">TB virtual interfaces -> agents</text><text class="small" x="745" y="328" text-anchor="middle">monitor AP -> scoreboard imp</text><text class="small" x="745" y="349" text-anchor="middle">rx_cfg -> sequence/monitor/scb</text>
      <path class="arrow-orange" d="M145 165 V235"/><path class="arrow-orange" d="M255 295 H330"/><path class="arrow-orange" d="M550 295 H625"/>

      <text class="ink" x="35" y="405">C. Run phase: repeated per frame and per active connection</text>
      <rect class="blue" x="35" y="425" width="190" height="135" rx="8"/><text class="ink" x="130" y="453" text-anchor="middle">base_vseq</text><text class="small" x="130" y="476" text-anchor="middle">select by CONNECT_NUM</text><text class="small" x="130" y="497" text-anchor="middle">fork ISPTX sequences</text><text class="small" x="130" y="518" text-anchor="middle">raise/drop objection</text><text class="small" x="130" y="539" text-anchor="middle">2/3 branch caveat</text>
      <rect class="purple" x="270" y="425" width="190" height="135" rx="8"/><text class="ink" x="365" y="453" text-anchor="middle">Frame preparation</text><text class="small" x="365" y="476" text-anchor="middle">process_cfg(frame)</text><text class="small" x="365" y="497" text-anchor="middle">get_reg() setting payload</text><text class="small" x="365" y="518" text-anchor="middle">read pattern PPM</text><text class="small" x="365" y="539" text-anchor="middle">run golden commands</text>
      <rect class="green" x="505" y="425" width="190" height="135" rx="8"/><text class="ink" x="600" y="453" text-anchor="middle">Stimulus and DUT</text><text class="small" x="600" y="476" text-anchor="middle">setting packets</text><text class="small" x="600" y="497" text-anchor="middle">pixel / blank packets</text><text class="small" x="600" y="518" text-anchor="middle">I2C or directed force</text><text class="small" x="600" y="539" text-anchor="middle">DUT transforms data</text>
      <rect class="blue" x="740" y="425" width="190" height="135" rx="8"/><text class="ink" x="835" y="453" text-anchor="middle">Monitor sampling</text><text class="small" x="835" y="476" text-anchor="middle">interface/hierarchy taps</text><text class="small" x="835" y="497" text-anchor="middle">transaction streams</text><text class="small" x="835" y="518" text-anchor="middle">actual PPM/text dumps</text><text class="small" x="835" y="539" text-anchor="middle">gated where implemented</text>
      <rect class="amber" x="975" y="425" width="190" height="135" rx="8"/><text class="ink" x="1070" y="453" text-anchor="middle">Scoreboard compare</text><text class="small" x="1070" y="476" text-anchor="middle">load golden by frame/id</text><text class="small" x="1070" y="497" text-anchor="middle">queue and align streams</text><text class="small" x="1070" y="518" text-anchor="middle">frame/line/pixel checks</text><text class="small" x="1070" y="539" text-anchor="middle">UVM error on mismatch</text>
      <path class="arrow" d="M225 492 H270"/><path class="arrow" d="M460 492 H505"/><path class="arrow-green" d="M695 492 H740"/><path class="arrow" d="M930 492 H975"/>
      <path class="arrow-purple dash" d="M1070 560 V600 H365 V560"/><text class="small" x="665" y="592">next frame: configuration, golden, stimulus and comparison advance together</text>

      <text class="ink" x="35" y="650">D. Completion and result</text>
      <rect class="box" x="35" y="670" width="270" height="90" rx="8"/><text class="ink" x="170" y="698" text-anchor="middle">Artifacts</text><text class="small" x="170" y="721" text-anchor="middle">outResult / actual PPM / logs</text><text class="small" x="170" y="742" text-anchor="middle">seed + waveform + command line</text>
      <rect class="purple" x="390" y="670" width="270" height="90" rx="8"/><text class="ink" x="525" y="698" text-anchor="middle">base_test.report_phase</text><text class="small" x="525" y="721" text-anchor="middle">count UVM_ERROR</text><text class="small" x="525" y="742" text-anchor="middle">print final pass/fail status</text>
      <rect class="green" x="745" y="670" width="270" height="90" rx="8"/><text class="ink" x="880" y="698" text-anchor="middle">Regression collection</text><text class="small" x="880" y="721" text-anchor="middle">case result + coverage</text><text class="small" x="880" y="742" text-anchor="middle">retain failing evidence</text>
      <path class="arrow-purple dash" d="M1070 560 V625 H170 V670"/><path class="arrow-purple" d="M305 715 H390"/><path class="arrow-purple" d="M660 715 H745"/>
      <g transform="translate(35 800)"><path class="arrow-orange" d="M0 8 H28"/><text class="small" x="36" y="12">build/config control</text><path class="arrow" d="M220 8 H248"/><text class="small" x="256" y="12">stimulus/transaction</text><path class="arrow-green" d="M465 8 H493"/><text class="small" x="501" y="12">sampled DUT data</text><path class="arrow-purple dash" d="M690 8 H718"/><text class="small" x="726" y="12">frame/result control</text></g>
    """
    checkers = """
      <text class="ink" x="30" y="34">DUT/interface tap</text><text class="ink" x="265" y="34">Monitor and emitted analysis ports</text><text class="ink" x="665" y="34">Scoreboard implementation</text><text class="ink" x="1010" y="34">Expected data / comparison</text>

      <rect class="green" x="30" y="60" width="180" height="105" rx="8"/><text class="ink" x="120" y="88" text-anchor="middle">data_merge_intf</text><text class="small" x="120" y="111" text-anchor="middle">plus analog_data_output_if</text><text class="small" x="120" y="132" text-anchor="middle">two output streams</text><text class="small" x="120" y="153" text-anchor="middle">rx_cfg</text>
      <rect class="blue" x="265" y="60" width="330" height="105" rx="8"/><text class="ink" x="430" y="88" text-anchor="middle">data_merge_monitor</text><text class="small" x="430" y="111" text-anchor="middle">out_data_dump_aport</text><text class="small" x="430" y="132" text-anchor="middle">out_data_dump_aport_1</text><text class="small" x="430" y="153" text-anchor="middle">writes data_merge_tr</text>
      <rect class="purple" x="665" y="60" width="280" height="105" rx="8"/><text class="ink" x="805" y="88" text-anchor="middle">data_merge_scoreboard</text><text class="small" x="805" y="111" text-anchor="middle">out_data_dump_aexport[_1]</text><text class="small" x="805" y="132" text-anchor="middle">writes actual PPM</text><text class="small" x="805" y="153" text-anchor="middle">gate: data_merge_check_on</text>
      <rect class="amber" x="1010" y="60" width="330" height="105" rx="8"/><text class="ink" x="1175" y="88" text-anchor="middle">Data Merge golden</text><text class="small" x="1175" y="111" text-anchor="middle">outResult[/id]/data_merge</text><text class="small" x="1175" y="132" text-anchor="middle">frame_N_data_merge&lt;id&gt;.txt</text><text class="small" x="1175" y="153" text-anchor="middle">pixel compare + UVM error</text>
      <path class="arrow-green" d="M210 112 H265"/><path class="arrow" d="M595 100 H665"/><path class="arrow" d="M595 137 H665"/><path class="arrow-green" d="M945 112 H1010"/>

      <rect class="green" x="30" y="205" width="180" height="120" rx="8"/><text class="ink" x="120" y="233" text-anchor="middle">digital_top_intf</text><text class="small" x="120" y="256" text-anchor="middle">plus analog_data_output_if</text><text class="small" x="120" y="277" text-anchor="middle">left/right odd/even</text><text class="small" x="120" y="298" text-anchor="middle">rx_cfg</text>
      <rect class="blue" x="265" y="205" width="330" height="120" rx="8"/><text class="ink" x="430" y="233" text-anchor="middle">digital_top_monitor</text><text class="small" x="430" y="256" text-anchor="middle">digital_top_ol / el_aport</text><text class="small" x="430" y="277" text-anchor="middle">digital_top_or / er_aport</text><text class="small" x="430" y="298" text-anchor="middle">four digital_top_tr streams</text><text class="small" x="430" y="319" text-anchor="middle">source line 31 uses data_merge_check_on</text>
      <rect class="purple" x="665" y="205" width="280" height="120" rx="8"/><text class="ink" x="805" y="233" text-anchor="middle">digital_top_scoreboard</text><text class="small" x="805" y="256" text-anchor="middle">OL / EL / OR / ER exports</text><text class="small" x="805" y="277" text-anchor="middle">four queues + actual PPM</text><text class="small" x="805" y="298" text-anchor="middle">gate: digital_top_check_on</text><text class="small" x="805" y="319" text-anchor="middle">optional DPLC PPM compare</text>
      <rect class="amber" x="1010" y="205" width="330" height="120" rx="8"/><text class="ink" x="1175" y="233" text-anchor="middle">Digital Top golden</text><text class="small" x="1175" y="256" text-anchor="middle">outResult/digital</text><text class="small" x="1175" y="277" text-anchor="middle">Lodd / Leven / Rodd / Reven</text><text class="small" x="1175" y="298" text-anchor="middle">DPLC_frame&lt;N&gt;_id&lt;id&gt;.ppm</text><text class="small" x="1175" y="319" text-anchor="middle">frame/line/pixel compare</text>
      <path class="arrow-green" d="M210 265 H265"/><path class="arrow" d="M595 238 H665"/><path class="arrow" d="M595 256 H665"/><path class="arrow" d="M595 274 H665"/><path class="arrow" d="M595 292 H665"/><path class="arrow-green" d="M945 265 H1010"/>

      <rect class="green" x="30" y="365" width="180" height="105" rx="8"/><text class="ink" x="120" y="393" text-anchor="middle">chopper_intf</text><text class="small" x="120" y="416" text-anchor="middle">analog_data_output_if</text><text class="small" x="120" y="437" text-anchor="middle">chopper dump tap</text><text class="small" x="120" y="458" text-anchor="middle">rx_cfg</text>
      <rect class="blue" x="265" y="365" width="330" height="105" rx="8"/><text class="ink" x="430" y="393" text-anchor="middle">chopper_monitor</text><text class="small" x="430" y="416" text-anchor="middle">chopper_dump_aport_d connected</text><text class="small" x="430" y="437" text-anchor="middle">aport_g declared, not connected</text><text class="small" x="430" y="458" text-anchor="middle">writes chopper_tr</text>
      <rect class="purple" x="665" y="365" width="280" height="105" rx="8"/><text class="ink" x="805" y="393" text-anchor="middle">chopper_scoreboard</text><text class="small" x="805" y="416" text-anchor="middle">out_data_chop_aexport_d</text><text class="small" x="805" y="437" text-anchor="middle">gate: chopper_check_on</text><text class="small" x="805" y="458" text-anchor="middle">chopper data compare</text>
      <rect class="amber" x="1010" y="365" width="330" height="105" rx="8"/><text class="ink" x="1175" y="393" text-anchor="middle">Chopper expectation</text><text class="small" x="1175" y="416" text-anchor="middle">model/config-derived stream</text><text class="small" x="1175" y="437" text-anchor="middle">transaction compare</text><text class="small" x="1175" y="458" text-anchor="middle">UVM error on mismatch</text>
      <path class="arrow-green" d="M210 417 H265"/><path class="arrow" d="M595 417 H665"/><path class="arrow-green" d="M945 417 H1010"/>

      <rect class="green" x="30" y="510" width="180" height="125" rx="8"/><text class="ink" x="120" y="538" text-anchor="middle">analog_data_output_if</text><text class="small" x="120" y="561" text-anchor="middle">pixel / POL / chopper</text><text class="small" x="120" y="582" text-anchor="middle">unlock / VBK</text><text class="small" x="120" y="603" text-anchor="middle">rx_cfg</text>
      <rect class="blue" x="265" y="510" width="330" height="125" rx="8"/><text class="ink" x="430" y="538" text-anchor="middle">analog_data_output_monitor</text><text class="small" x="430" y="561" text-anchor="middle">mon / pol / chopper ports</text><text class="small" x="430" y="582" text-anchor="middle">unlock / vbk ports</text><text class="small" x="430" y="603" text-anchor="middle">5 transaction streams</text><text class="small" x="430" y="624" text-anchor="middle">main task gate: analog_check_on</text>
      <rect class="purple" x="665" y="510" width="280" height="125" rx="8"/><text class="ink" x="805" y="538" text-anchor="middle">analog_data_output_scoreboard</text><text class="small" x="805" y="561" text-anchor="middle">monitor / pol / chopper imps</text><text class="small" x="805" y="582" text-anchor="middle">unlock / vbk imps</text><text class="small" x="805" y="603" text-anchor="middle">pixel + control checks</text><text class="small" x="805" y="624" text-anchor="middle">id = env_cfg.id</text>
      <rect class="amber" x="1010" y="510" width="330" height="125" rx="8"/><text class="ink" x="1175" y="538" text-anchor="middle">Analog/control expectation</text><text class="small" x="1175" y="561" text-anchor="middle">pixel sequence and control timing</text><text class="small" x="1175" y="582" text-anchor="middle">POL / chop / unlock / VBK</text><text class="small" x="1175" y="603" text-anchor="middle">per-event checks</text><text class="small" x="1175" y="624" text-anchor="middle">UVM error on mismatch</text>
      <path class="arrow-green" d="M210 572 H265"/><path class="arrow" d="M595 540 H665"/><path class="arrow" d="M595 556 H665"/><path class="arrow" d="M595 572 H665"/><path class="arrow" d="M595 588 H665"/><path class="arrow" d="M595 604 H665"/><path class="arrow-green" d="M945 572 H1010"/>

      <rect class="green" x="30" y="675" width="180" height="125" rx="8"/><text class="ink" x="120" y="703" text-anchor="middle">analog_data_output_if</text><text class="small" x="120" y="726" text-anchor="middle">DRD input tap</text><text class="small" x="120" y="747" text-anchor="middle">DRD output tap</text><text class="small" x="120" y="768" text-anchor="middle">rx_cfg</text>
      <rect class="blue" x="265" y="675" width="330" height="125" rx="8"/><text class="ink" x="430" y="703" text-anchor="middle">DRD input + output monitors</text><text class="small" x="430" y="726" text-anchor="middle">each has out_aport</text><text class="small" x="430" y="747" text-anchor="middle">independent transaction types</text><text class="small" x="430" y="768" text-anchor="middle">gates: drd_input/output_check_on</text>
      <rect class="purple" x="665" y="675" width="280" height="125" rx="8"/><text class="ink" x="805" y="703" text-anchor="middle">DRD input/output scoreboards</text><text class="small" x="805" y="726" text-anchor="middle">drdod_imp per path</text><text class="small" x="805" y="747" text-anchor="middle">input pattern validation</text><text class="small" x="805" y="768" text-anchor="middle">output/bypass validation</text>
      <rect class="amber" x="1010" y="675" width="330" height="125" rx="8"/><text class="ink" x="1175" y="703" text-anchor="middle">DRD protocol expectation</text><text class="small" x="1175" y="726" text-anchor="middle">DRD_PANEL / cyclic / OD params</text><text class="small" x="1175" y="747" text-anchor="middle">DRDOD_EN and bypass mapping</text><text class="small" x="1175" y="768" text-anchor="middle">input and output checked separately</text>
      <path class="arrow-green" d="M210 737 H265"/><path class="arrow" d="M595 718 H665"/><path class="arrow" d="M595 758 H665"/><path class="arrow-green" d="M945 737 H1010"/>

      <rect class="box" x="30" y="835" width="1310" height="48" rx="8"/><text class="small" x="685" y="855" text-anchor="middle">checker_agent.build_phase creates every monitor and scoreboard unconditionally; checker flags gate sampling/comparison inside components.</text><text class="small" x="685" y="875" text-anchor="middle">Connections above are exact checker_agent.connect_phase analysis_port -> analysis_imp/export mappings.</text>
      <g transform="translate(40 915)"><path class="arrow-green" d="M0 8 H28"/><text class="small" x="36" y="12">sampled DUT/interface data</text><path class="arrow" d="M255 8 H283"/><text class="small" x="291" y="12">analysis transaction</text><path class="arrow-green" d="M485 8 H513"/><text class="small" x="521" y="12">expected/compare path</text></g>
    """
    lifecycle = """
      <rect class="blue" x="30" y="72" width="150" height="76" rx="5"/><text class="ink" x="105" y="101" text-anchor="middle">选择模板</text><text class="small" x="105" y="123" text-anchor="middle">closest testcase</text>
      <rect class="box" x="220" y="72" width="150" height="76" rx="5"/><text class="ink" x="295" y="101" text-anchor="middle">创建目录</text><text class="small" x="295" y="123" text-anchor="middle">create_new_case.sh</text>
      <rect class="amber" x="410" y="72" width="150" height="76" rx="5"/><text class="ink" x="485" y="101" text-anchor="middle">配置证据</text><text class="small" x="485" y="123" text-anchor="middle">macro / cfg / pattern</text>
      <rect class="green" x="600" y="72" width="150" height="76" rx="5"/><text class="ink" x="675" y="101" text-anchor="middle">启用 checker</text><text class="small" x="675" y="123" text-anchor="middle">*_check_on</text>
      <rect class="blue" x="790" y="72" width="150" height="76" rx="5"/><text class="ink" x="865" y="101" text-anchor="middle">运行与定位</text><text class="small" x="865" y="123" text-anchor="middle">run_tc.sh / artifacts</text>
      <path class="arrow" d="M180 110 H220"/><path class="arrow" d="M370 110 H410"/><path class="arrow" d="M560 110 H600"/><path class="arrow" d="M750 110 H790"/>
      <text class="small" x="31" y="190">Only add to case_list.txt after the testcase has deterministic checks and reproducible inputs.</text>
    """
    frame_sequence = """
      <text class="ink" x="70" y="38">Testcase</text><text class="ink" x="270" y="38">env_cfg</text><text class="ink" x="470" y="38">Golden model</text><text class="ink" x="670" y="38">ISPTX / DUT</text><text class="ink" x="870" y="38">Scoreboard</text>
      <path class="dash" d="M90 55 V430 M290 55 V430 M490 55 V430 M690 55 V430 M890 55 V430" stroke="#d1d5db"/>
      <path class="arrow-orange" d="M90 88 H280"/><text class="small" x="146" y="80">user_def + cfg_frameN</text>
      <path class="arrow-green" d="M290 135 H480"/><text class="small" x="350" y="127">model arguments</text>
      <path class="arrow" d="M290 190 H680"/><text class="small" x="414" y="182">get_reg() / setting packet</text>
      <path class="arrow-green" d="M490 240 H880"/><text class="small" x="615" y="232">golden PPM / text</text>
      <path class="arrow" d="M690 290 C790 290 790 315 880 315"/><text class="small" x="735" y="282">monitor transaction</text>
      <path class="arrow-purple dash" d="M890 355 H700"/><text class="small" x="745" y="347">pass / UVM error</text>
      <rect class="blue" x="42" y="385" width="896" height="36" rx="8"/><text class="small" x="490" y="408" text-anchor="middle">Repeat for frame_idx: configuration, golden, stimulus and compare must use the same frame context.</text>
      <g transform="translate(30 450)"><path class="arrow" d="M0 8 H28"/><text class="small" x="36" y="12">packet/data</text><path class="arrow-green" d="M145 8 H173"/><text class="small" x="181" y="12">golden artifact</text><path class="arrow-orange" d="M325 8 H353"/><text class="small" x="361" y="12">configuration</text><path class="arrow-purple dash" d="M485 8 H513"/><text class="small" x="521" y="12">result/control</text></g>
    """
    debug_flow = """
      <rect class="amber" x="40" y="40" width="190" height="70" rx="8"/><text class="ink" x="135" y="69" text-anchor="middle">UVM error / mismatch</text><text class="small" x="135" y="91" text-anchor="middle">capture first failing frame</text>
      <rect class="blue" x="300" y="40" width="190" height="70" rx="8"/><text class="ink" x="395" y="69" text-anchor="middle">Checker enabled?</text><text class="small" x="395" y="91" text-anchor="middle">inspect *_check_on</text>
      <rect class="green" x="560" y="40" width="190" height="70" rx="8"/><text class="ink" x="655" y="69" text-anchor="middle">Golden exists?</text><text class="small" x="655" y="91" text-anchor="middle">command + output path</text>
      <rect class="purple" x="820" y="40" width="190" height="70" rx="8"/><text class="ink" x="915" y="69" text-anchor="middle">Earliest bad layer</text><text class="small" x="915" y="91" text-anchor="middle">merge / digital / analog</text>
      <path class="arrow-orange" d="M230 75 H300"/><path class="arrow-orange" d="M490 75 H560"/><path class="arrow-orange" d="M750 75 H820"/>
      <rect class="box" x="300" y="175" width="190" height="74" rx="8"/><text class="ink" x="395" y="204" text-anchor="middle">Configuration problem</text><text class="small" x="395" y="226" text-anchor="middle">macro / cfg / frame index</text>
      <rect class="box" x="560" y="175" width="190" height="74" rx="8"/><text class="ink" x="655" y="204" text-anchor="middle">Model / file problem</text><text class="small" x="655" y="226" text-anchor="middle">arguments / stale artifact</text>
      <rect class="box" x="820" y="175" width="190" height="74" rx="8"/><text class="ink" x="915" y="204" text-anchor="middle">RTL / protocol problem</text><text class="small" x="915" y="226" text-anchor="middle">frame / line / pixel waveform</text>
      <path class="arrow-purple dash" d="M395 110 V175"/><text class="small" x="406" y="146">off / wrong target</text><path class="arrow-purple dash" d="M655 110 V175"/><text class="small" x="666" y="146">missing / stale</text><path class="arrow-purple dash" d="M915 110 V175"/><text class="small" x="926" y="146">localized</text>
      <rect class="blue" x="300" y="310" width="710" height="55" rx="8"/><text class="ink" x="655" y="337" text-anchor="middle">Re-run one deterministic case and preserve log, golden, actual dump and seed</text><text class="small" x="655" y="356" text-anchor="middle">Only return to regression after the failure can be reproduced and the checker catches it.</text>
      <path class="arrow" d="M395 249 V282 H655 V310"/><path class="arrow" d="M655 249 V310"/><path class="arrow" d="M915 249 V282 H655 V310"/>
      <g transform="translate(40 402)"><path class="arrow-orange" d="M0 8 H28"/><text class="small" x="36" y="12">triage sequence</text><path class="arrow-purple dash" d="M180 8 H208"/><text class="small" x="216" y="12">decision branch</text><path class="arrow" d="M370 8 H398"/><text class="small" x="406" y="12">fix and reproduce</text></g>
    """
    isptx_driver_flow = """
      <rect class="blue" x="30" y="70" width="170" height="105" rx="8"/><text class="ink" x="115" y="100" text-anchor="middle">base_vseq</text><text class="small" x="115" y="125" text-anchor="middle">start sequence on sqrN</text><text class="small" x="115" y="147" text-anchor="middle">CONNECT_NUM</text>
      <rect class="purple" x="245" y="45" width="220" height="155" rx="8"/><text class="ink" x="355" y="75" text-anchor="middle">isptx_sequence</text><text class="small" x="355" y="100" text-anchor="middle">TRAIN_SEND / ACTIVE</text><text class="small" x="355" y="122" text-anchor="middle">setting + blank + pixel</text><text class="small" x="355" y="144" text-anchor="middle">small typed transactions</text><text class="small" x="355" y="166" text-anchor="middle">unlock -> retrain</text>
      <rect class="box" x="510" y="70" width="165" height="105" rx="8"/><text class="ink" x="592" y="100" text-anchor="middle">sequencer</text><text class="small" x="592" y="125" text-anchor="middle">request arbitration</text><text class="small" x="592" y="147" text-anchor="middle">seq_item_export</text>
      <rect class="amber" x="720" y="45" width="220" height="155" rx="8"/><text class="ink" x="830" y="75" text-anchor="middle">isptx_driver</text><text class="small" x="830" y="100" text-anchor="middle">get_next_item / item_done</text><text class="small" x="830" y="122" text-anchor="middle">dispatch by id</text><text class="small" x="830" y="144" text-anchor="middle">9 clocks per symbol</text><text class="small" x="830" y="166" text-anchor="middle">lane0/lane1 parallel</text>
      <rect class="green" x="985" y="70" width="175" height="105" rx="8"/><text class="ink" x="1072" y="100" text-anchor="middle">DUT link input</text><text class="small" x="1072" y="125" text-anchor="middle">rxp0/rxn0</text><text class="small" x="1072" y="147" text-anchor="middle">rxp1/rxn1</text>
      <path class="arrow" d="M200 122 H245 M465 122 H510 M675 122 H720 M940 122 H985"/>
      <rect class="box" x="245" y="260" width="695" height="80" rx="8"/><text class="ink" x="592" y="288" text-anchor="middle">Frame protocol</text><text class="small" x="592" y="313" text-anchor="middle">VBP setting -> HBP -> BAC+POL -> pixel -> EOL -> HFP -> VFP</text><text class="small" x="592" y="333" text-anchor="middle">LANE / TRAINING / REGISTER / HBK / BAC / POL / EOL / PIXEL / DUMMY</text>
      <path class="arrow-orange dash" d="M355 200 V260 M830 200 V260"/>
    """
    checker_position = """
      <rect class="blue" x="25" y="75" width="150" height="90" rx="8"/><text class="ink" x="100" y="105" text-anchor="middle">RX/decode</text><text class="small" x="100" y="132" text-anchor="middle">DUT input</text>
      <rect class="green" x="215" y="55" width="170" height="130" rx="8"/><text class="ink" x="300" y="87" text-anchor="middle">Data Merge</text><text class="small" x="300" y="114" text-anchor="middle">2 streams</text><text class="small" x="300" y="136" text-anchor="middle">merge/mapping</text><text class="small" x="300" y="158" text-anchor="middle">monitor -> scb</text>
      <rect class="green" x="425" y="55" width="170" height="130" rx="8"/><text class="ink" x="510" y="87" text-anchor="middle">Digital Top</text><text class="small" x="510" y="114" text-anchor="middle">OL/EL/OR/ER</text><text class="small" x="510" y="136" text-anchor="middle">4 streams</text><text class="small" x="510" y="158" text-anchor="middle">monitor -> scb</text>
      <rect class="green" x="635" y="55" width="170" height="130" rx="8"/><text class="ink" x="720" y="87" text-anchor="middle">Chopper</text><text class="small" x="720" y="114" text-anchor="middle">D port connected</text><text class="small" x="720" y="136" text-anchor="middle">G port unconnected</text><text class="small" x="720" y="158" text-anchor="middle">monitor -> scb</text>
      <rect class="green" x="845" y="55" width="190" height="130" rx="8"/><text class="ink" x="940" y="87" text-anchor="middle">Analog/control</text><text class="small" x="940" y="114" text-anchor="middle">pixel/POL/chop</text><text class="small" x="940" y="136" text-anchor="middle">unlock/VBK</text><text class="small" x="940" y="158" text-anchor="middle">5 paths</text>
      <path class="arrow-green" d="M175 120 H215 M385 120 H425 M595 120 H635 M805 120 H845"/>
      <rect class="purple" x="425" y="270" width="170" height="95" rx="8"/><text class="ink" x="510" y="300" text-anchor="middle">DRD input</text><text class="small" x="510" y="327" text-anchor="middle">panel/input pattern</text><text class="small" x="510" y="349" text-anchor="middle">independent path</text>
      <rect class="purple" x="680" y="270" width="170" height="95" rx="8"/><text class="ink" x="765" y="300" text-anchor="middle">DRD output</text><text class="small" x="765" y="327" text-anchor="middle">output/bypass</text><text class="small" x="765" y="349" text-anchor="middle">independent path</text>
      <path class="arrow-purple" d="M510 185 V270 M595 317 H680 M765 270 V220 H940 V185"/>
      <text class="small" x="530" y="420" text-anchor="middle">Parallel DUT taps: debug by the earliest failing stage.</text>
    """
    golden_flow = """
      <rect class="blue" x="30" y="125" width="180" height="105" rx="8"/><text class="ink" x="120" y="155" text-anchor="middle">Frame inputs</text><text class="small" x="120" y="180" text-anchor="middle">pattern/frame_N.ppm</text><text class="small" x="120" y="202" text-anchor="middle">rx_cfg / frame / id</text>
      <rect class="purple" x="270" y="35" width="220" height="80" rx="8"/><text class="ink" x="380" y="65" text-anchor="middle">pic_process</text><text class="small" x="380" y="90" text-anchor="middle">outResult text</text>
      <rect class="purple" x="270" y="145" width="220" height="80" rx="8"/><text class="ink" x="380" y="175" text-anchor="middle">dplc.pl</text><text class="small" x="380" y="200" text-anchor="middle">DPLC_frameN_idM.ppm</text>
      <rect class="purple" x="270" y="255" width="220" height="80" rx="8"/><text class="ink" x="380" y="285" text-anchor="middle">drdod_process.py</text><text class="small" x="380" y="310" text-anchor="middle">drdod_out_frameN.ppm</text>
      <path class="arrow" d="M210 160 H240 V75 H270 M210 177 H270 M210 195 H240 V295 H270"/>
      <rect class="green" x="555" y="75" width="235" height="220" rx="8"/><text class="ink" x="672" y="108" text-anchor="middle">Scoreboard loader</text><text class="small" x="672" y="140" text-anchor="middle">OD_k -> DRDOD PPM</text><text class="small" x="672" y="166" text-anchor="middle">DPLC_MODE -> DPLC PPM</text><text class="small" x="672" y="192" text-anchor="middle">else -> outResult text</text><text class="small" x="672" y="230" text-anchor="middle">frame/id + channel/SHL</text><text class="small" x="672" y="256" text-anchor="middle">build model queues</text><text class="small" x="672" y="278" text-anchor="middle">no C API call</text>
      <path class="arrow-green" d="M490 75 H555 M490 185 H530 V160 H555 M490 295 H530 V245 H555"/>
      <rect class="amber" x="850" y="75" width="175" height="95" rx="8"/><text class="ink" x="937" y="105" text-anchor="middle">Monitor actual</text><text class="small" x="937" y="132" text-anchor="middle">DUT queues / PPM</text>
      <rect class="box" x="850" y="225" width="175" height="95" rx="8"/><text class="ink" x="937" y="255" text-anchor="middle">Compare</text><text class="small" x="937" y="282" text-anchor="middle">frame/line/subpixel</text><text class="small" x="937" y="304" text-anchor="middle">uvm_error</text>
      <path class="arrow" d="M790 185 H820 V272 H850 M937 170 V225"/>
    """
    special_modes = """
      <rect class="blue" x="30" y="65" width="170" height="90" rx="8"/><text class="ink" x="115" y="95" text-anchor="middle">Mode select</text><text class="small" x="115" y="122" text-anchor="middle">DPLC_MODE / OD_k</text>
      <rect class="purple" x="260" y="35" width="245" height="160" rx="8"/><text class="ink" x="382" y="65" text-anchor="middle">DPLC</text><text class="small" x="382" y="92" text-anchor="middle">load PPM + subpix adjust</text><text class="small" x="382" y="114" text-anchor="middle">011/2-lane remap</text><text class="small" x="382" y="136" text-anchor="middle">interleaved RGB golden</text><text class="small" x="382" y="158" text-anchor="middle">120-point offset branch</text><text class="small" x="382" y="180" text-anchor="middle">VACT compare</text>
      <rect class="purple" x="570" y="35" width="245" height="160" rx="8"/><text class="ink" x="692" y="65" text-anchor="middle">DRDOD</text><text class="small" x="692" y="92" text-anchor="middle">direct 8-bit generation</text><text class="small" x="692" y="114" text-anchor="middle">panel/cyclic/nomatch</text><text class="small" x="692" y="136" text-anchor="middle">8B9B to DUT</text><text class="small" x="692" y="158" text-anchor="middle">DRDOD golden PPM</text><text class="small" x="692" y="180" text-anchor="middle">VACT*2 compare</text>
      <path class="arrow" d="M200 100 H260 M200 125 H230 V115 H570"/>
      <rect class="green" x="260" y="270" width="245" height="95" rx="8"/><text class="ink" x="382" y="300" text-anchor="middle">Analog DPLC branch</text><text class="small" x="382" y="327" text-anchor="middle">channel fill + SHL handling</text>
      <rect class="green" x="570" y="245" width="245" height="145" rx="8"/><text class="ink" x="692" y="275" text-anchor="middle">DRD checks</text><text class="small" x="692" y="302" text-anchor="middle">Analog PPM pixels</text><text class="small" x="692" y="324" text-anchor="middle">DRD input pattern</text><text class="small" x="692" y="346" text-anchor="middle">DRD output/bypass</text><text class="small" x="692" y="368" text-anchor="middle">independent paths</text>
      <path class="arrow-green" d="M382 195 V270 M692 195 V245"/>
    """
    diagrams = {
        "architecture.svg": svg_document(1200, 720, architecture, "HV2M23 UVM verification environment architecture"),
        "verification-flow.svg": svg_document(1200, 840, flow, "HV2M23 end-to-end verification execution flow"),
        "checker-flow.svg": svg_document(1380, 960, checkers, "HV2M23 checker architecture"),
        "case-lifecycle.svg": svg_document(970, 220, lifecycle, "HV2M23 testcase lifecycle"),
        "frame-sequence.svg": svg_document(980, 490, frame_sequence, "HV2M23 per-frame verification sequence"),
        "debug-flow.svg": svg_document(1050, 450, debug_flow, "HV2M23 checker failure triage flow"),
        "isptx-driver-flow.svg": svg_document(1200, 380, isptx_driver_flow, "ISPTX sequence to DUT driver flow"),
        "checker-position.svg": svg_document(1060, 450, checker_position, "Checker positions in the DUT data path"),
        "golden-flow.svg": svg_document(1060, 370, golden_flow, "C model golden generation and checker loading"),
        "special-modes.svg": svg_document(850, 420, special_modes, "DPLC and DRDOD special processing"),
    }
    for name, content in diagrams.items():
        (ASSET_DIR / name).write_text(content, encoding="utf-8")


def nav(active: str) -> str:
    """Render the global navigation."""
    links = "".join(
        f'<a href="{href}" class="{"active" if href == active else ""}">{label}</a>'
        for href, label in NAV_ITEMS
    )
    return f'<header class="topbar"><div class="topbar-inner"><a class="brand" href="index.html"><span>HV2M23</span> DV Guide</a><nav class="nav">{links}</nav></div></header>'


def page(active: str, title: str, lead: str, body: str) -> str:
    """Render one complete guide page."""
    section_links = []
    section_index = 0

    def add_section_id(match: re.Match[str]) -> str:
        nonlocal section_index
        section_index += 1
        attributes, heading = match.group(1), match.group(2)
        section_id = f"section-{section_index}"
        label = re.sub(r"<[^>]+>", "", heading)
        section_links.append((section_id, label))
        return f'<h2 id="{section_id}"{attributes}>{heading}</h2>'

    body = re.sub(r"<h2([^>]*)>(.*?)</h2>", add_section_id, body,
                  flags=re.DOTALL)
    section_nav = "".join(
        f'<a href="#{section_id}">{esc(label)}</a>'
        for section_id, label in section_links)
    return dedent(f"""\
        <!doctype html>
        <!-- Auto-generated by tools/build_guide.py -->
        <html lang="zh-CN">
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
          <title>{esc(title)} - HV2M23 DV Guide</title>
          <link rel="stylesheet" href="assets/guide.css?v=20260727">
        </head>
        <body>
          {nav(active)}
          <main class="layout">
            <header class="page-head">
              <div class="eyebrow">HV2M23 / DV_TCON_C</div>
              <h1>{title}</h1>
              <p class="lead">{lead}</p>
              <div class="source-stamp">Source baseline: E:\\DV_TCON_C · audited {BUILD_DATE}</div>
            </header>
            <div class="content-grid">
              <aside class="section-nav"><strong>本页目录</strong>{section_nav}</aside>
              <div class="page-content">{body}</div>
            </div>
            <footer class="footer">HV2M23 DV_TCON_C 验证环境使用指南 · 源码审计版 {BUILD_DATE}</footer>
          </main>
        </body>
        </html>
    """)


def figure(filename: str, caption: str) -> str:
    """Render a diagram figure."""
    return f'<figure class="figure"><img src="assets/{filename}" alt="{esc(caption)}"><figcaption class="caption">{caption}</figcaption></figure>'


def build_pages(
        cases: list[dict[str, object]], metadata: dict[str, object]) -> dict[str, str]:
    """Build all guide pages from audited source facts."""
    def cell_html(value: object) -> str:
        return esc(value).replace("\n", "<br>")

    passed_count = sum(
        1 for item in cases if str(item["status"]).upper() == "PASS")
    linked_count = sum(1 for item in cases if item["linkedSource"])
    excel_only_count = len(cases) - linked_count
    category_count = len({str(item["category"]) for item in cases})
    index_body = f"""
      <div class="stats"><div class="stat"><strong>{len(cases)}</strong><span>Excel named plan rows</span></div><div class="stat"><strong>{passed_count}</strong><span>Excel PASS rows</span></div><div class="stat"><strong>{linked_count}</strong><span>rows linked to source</span></div><div class="stat"><strong>{excel_only_count}</strong><span>Excel-only rows</span></div></div>
      <div class="note"><strong>Case 事实源：</strong>具体 case 范围、验证目标、check 项、owner 和状态以 <code>3.HV2M23_EDA_case_list.xlsx</code> 为准；tests 源码只作为实现证据和补充，不再反向扩大计划范围。</div>
      {figure('architecture.svg', '验证环境总览：1-4 路 ISPTX/I2C agent 驱动 DUT，多个 monitor/scoreboard 从不同层级闭环检查。')}
      <h2>从哪里开始</h2>
      <div class="grid-3">
        <section class="panel"><h3>理解数据流</h3><p>先看概览、TB 架构与激励页，确认 testcase 配置如何变成寄存器包、pixel stream 和 golden 文件。</p><a href="overview.html">打开概览</a></section>
        <section class="panel"><h3>定位检查失败</h3><p>按 checker 层级判断错误发生在 data merge、digital split、chopper/analog 还是 DRDOD。</p><a href="checkers.html">打开检查机制</a></section>
        <section class="panel"><h3>查具体 Case</h3><p>Case 索引以 Excel 计划为准，并叠加可关联的宏、checker 开关、force 和源码证据。</p><a href="cases.html">打开 Case 索引</a></section>
      </div>
      <div class="note"><strong>重要：</strong>用例名里的 CHSEL 是寄存器映射参数，不是编译宏 CHIP_SEL。必须以每个目录的 <code>user_def.sv</code> 为准。</div>
    """

    index_body += """
      <h2>指南覆盖范围</h2>
      <div class="grid-3">
        <section class="panel"><h3>环境搭建</h3><p>从 <code>chip_tb_top</code>、<code>base_test</code> 到多路 source env，解释组件是谁创建、interface 如何下发。</p></section>
        <section class="panel"><h3>激励闭环</h3><p>把宏、cfg frame、PPM、golden 命令、setting/pixel packet 和逐帧比较串成一条可追踪链路。</p></section>
        <section class="panel"><h3>Case 审计</h3><p>{len(cases)} 条 Excel 命名计划项保留描述、check 点、owner 和状态，同时显示源码关联差异。</p></section>
      </div>
      <h2>推荐阅读路径</h2>
      <ol class="steps"><li><strong>第一次接触环境：</strong>概览 -> TB 架构 -> 激励与 Golden。</li><li><strong>正在定位 fail：</strong>检查机制 -> 运行与回归 -> Case 计划索引。</li><li><strong>准备新增 case：</strong>先在 Excel 计划中建立目标和 check 点，再关联源码目录。</li><li><strong>准备下一项目：</strong>阅读复用与移植页，按平台层、适配层、内容层拆分。</li></ol>
    """

    overview_body = f"""
      {figure('verification-flow.svg', '端到端执行流程：从 case 选择、编译、UVM build/connect，到逐帧激励、checker 比较、report 和回归结果收集。')}
      <h2>事实源优先级</h2>
      <div class="grid-2">
        <section class="panel accent"><h3>1. Testcase 源码</h3><p><code>top/tests/&lt;case&gt;/</code> 中的 test class、<code>user_def.sv</code>、<code>cfg_frame*.txt</code> 和 pattern 是该 case 的最终事实。</p></section>
        <section class="panel"><h3>2. 环境实现</h3><p><code>base_test.sv</code>、<code>base_vseq.sv</code>、<code>env.sv</code>、<code>env_cfg.sv</code> 和 checker 源码决定实际行为。</p></section>
      </div>
      <h2>目录职责</h2>
      <table><tr><th>路径</th><th>职责</th></tr>
      <tr><td><code>top/tb/</code></td><td>顶层接口、DUT 连接、virtual interface 下发、base_test。</td></tr>
      <tr><td><code>top/vseq/</code></td><td>根据 CONNECT_NUM 选择并行 ISPTX sequence 和目标 sequencer；当前分支结构存在 2/3 路时额外执行默认 sequence 的风险。</td></tr>
      <tr><td><code>top/agents/isprx_env/</code></td><td>ISPTX/I2C agent、env_cfg、golden model 调用及 checker。</td></tr>
      <tr><td><code>top/tests/</code></td><td>每个 testcase 的宏、配置帧、pattern、定向 force 和回归清单。</td></tr>
      <tr><td><code>script/run_tc.sh</code></td><td>编译、随机仿真、Verdi 和 coverage 的统一入口。</td></tr></table>
      <h2>已修正的常见误解</h2>
      <ul><li><code>CHSEL</code> 与 <code>CHIP_SEL</code> 不是同一配置。</li><li>DRDOD checker 不只看 analog 图片；环境已有独立 input/output monitor 与 scoreboard。</li><li>Digital Top checker 分为 OL/OR/EL/ER 四路，而不是单一输出。</li><li>当前 DPLC 命令接口以 <code>isptx_sequence.sv</code> 中实际调用为准，不引用其他项目脚本的参数。</li></ul>
    """

    overview_body += """
      <h2>配置的四个层次</h2>
      <table><tr><th>层次</th><th>典型内容</th><th>何时生效</th><th>核对方式</th></tr>
      <tr><td>编译宏</td><td><code>PAIR_NUM</code>、<code>PORT_NUM</code>、<code>FRAME_NUM</code>、<code>CHIP_SEL</code></td><td>编译/elaboration</td><td>查看具体 case 的 <code>user_def.sv</code></td></tr>
      <tr><td>运行配置</td><td><code>rx_cfg</code>/<code>env_cfg</code> checker 开关和注入参数</td><td>build/connect/run phase</td><td>查看 testcase class 的赋值</td></tr>
      <tr><td>逐帧寄存器</td><td><code>cfg_frameN.txt</code></td><td>每帧 setting packet 前</td><td>对照 <code>process_cfg()</code> 与发送日志</td></tr>
      <tr><td>定向控制</td><td><code>force</code>、I2C transaction、协议异常注入</td><td>case 指定时刻</td><td>查看 testcase task 与波形</td></tr></table>
      <div class="note"><strong>审计原则：</strong>目录名只能用于检索，不能替代 <code>user_def.sv</code>、cfg 内容和 testcase 代码。出现冲突时，以实际执行代码为准。</div>
    """

    arch_body = f"""
      {figure('architecture.svg', 'UVM 组件关系图，依据 base_test.sv、env.sv 和 checker_agent.sv。')}
      <h2>组件与源码证据</h2>
      <table><tr><th>组件</th><th>实际职责</th><th>源码</th></tr>
      <tr><td>base_test</td><td>按 CONNECT_NUM 创建1到4个 source_driver_env，并连接 virtual sequencer 到各 ISPTX/I2C sequencer。</td><td><code>top/tb/base_test.sv</code></td></tr>
      <tr><td>source_driver_env</td><td>创建 isptx_agent、i2c_agent、checker_agent 和 env_cfg；向 monitor 分发 interface。</td><td><code>top/agents/isprx_env/env.sv</code></td></tr>
      <tr><td>base_vseq</td><td>使用 fork 启动与 CONNECT_NUM 对应的 sequence；但 2/3 路分支后会继续命中最后一个 if-else 的 else，按源码还会启动一次默认 sequence。</td><td><code>top/vseq/base_vseq.sv</code></td></tr>
      <tr><td>chip_tb_top</td><td>实例化4组 ISPTX/I2C interface，并通过 config_db 绑定到对应 env。</td><td><code>top/tb/chip_tb_top.sv</code></td></tr>
      <tr><td>checker_agent</td><td>集中创建并连接 Data Merge、Digital Top、Chopper、Analog、DRD input/output monitor 和 scoreboard。</td><td><code>top/agents/isprx_env/checker_env/checker_agent.sv</code></td></tr></table>
      <div class="note"><strong>三个数量参数：</strong><code>CONNECT_NUM</code> 控制 env、interface index 和 sequence 选择；<code>PAIR_NUM</code>、<code>PORT_NUM</code> 用于频率、像素拆分、PPM 宽度及模型参数。不能用 PAIR_NUM 代替 CONNECT_NUM。</div>
      <div class="note"><strong>源码审计发现：</strong><code>base_vseq.sv</code> 使用 <code>if(2)</code>、<code>if(3)</code>、<code>if(4)...else</code>，不是完整互斥链。CONNECT_NUM 为 2 或 3 时会在并行分支后再执行最后的默认单路 sequence。指南按实际代码记录；建议 RTL/DV 负责人确认是否应改为 <code>else if</code>。</div>
    """

    arch_body += """
      <h2>Build 与 Connect 的关键关系</h2>
      <div class="grid-2"><section class="panel"><h3>创建关系</h3><ul><li><code>base_test</code> 创建 virtual sequencer 和 source env。</li><li><code>source_driver_env</code> 创建 ISPTX、I2C 与 checker agent。</li><li><code>checker_agent</code> 依据配置创建所需 monitor/scoreboard。</li></ul></section><section class="panel"><h3>连接关系</h3><ul><li>virtual sequencer 保存各实体 sequencer 句柄。</li><li>monitor 的 analysis port 连接对应 scoreboard implementation。</li><li>interface 通过 config_db 从 TB top 下发到 agent/monitor。</li></ul></section></div>
      <h2>多连接扩展时要核对</h2><ul><li>每个 connection 的 virtual interface key 是否与 <code>sd_env_N</code> 对应。</li><li>virtual sequencer 的 ISPTX/I2C handle 是否全部连接。</li><li>实际启动的 sequence 次数是否符合 <code>CONNECT_NUM</code>，尤其检查 2/3 路的额外默认 sequence。</li><li>文件名、golden 输出和 checker 队列是否携带 <code>env_cfg.id</code>，避免多连接互相覆盖。</li><li><code>PAIR_NUM</code>/<code>PORT_NUM</code> 是数据布局参数，不是 env 实例数。</li></ul>
    """

    stimulus_body = f"""
      {figure('verification-flow.svg', '激励与 golden 共用同一份 env_cfg：寄存器配置、图像模型和 packet 发送在每帧对齐。')}
      <h2>每帧执行顺序</h2>
      <ol class="steps"><li><strong>读取 case 资产。</strong> <code>user_def.sv</code> 决定静态宏，<code>cfg_frameN.txt</code> 由 env_cfg 的 <code>process_cfg()</code> 解析。</li><li><strong>生成寄存器 payload。</strong> <code>get_reg()</code> 将当前 frame 配置打包，<code>send_register()</code> 发送 setting line。</li><li><strong>运行 golden。</strong> sequence 调用 <code>pic_process</code>；DRDOD 调用 <code>drdod_process.py</code>；DPLC 调用当前环境内的 <code>dplc.pl</code>。</li><li><strong>发送 pixel stream。</strong> 从 <code>pattern/frame_N.ppm</code> 读取 P3 数据，按协议插入 blank、setting 与异常注入。</li><li><strong>采样与比较。</strong> monitor 输出 transaction/PPM，scoreboard 按 frame 和输出通道加载 golden。</li></ol>
      {figure('isptx-driver-flow.svg', 'ISPTX sequence 到 DUT：sequence 负责协议和 transaction，driver 负责逐 bit 引脚时序。')}
      <h2>ISPTX Sequence 到 DUT 的具体链路</h2>
      <table><tr><th>阶段</th><th>实际动作</th><th>源码位置</th></tr>
      <tr><td>启动 sequence</td><td><code>base_vseq</code> 将 <code>isptx_sequence[_1..3]</code> 启动在对应 <code>p_isptx_sqrN</code>；<code>body()</code> 取得 rx_cfg、vif、SIM_TEST_PATH 和 FILE_PATH 后调用 <code>send_isp()</code>。</td><td><code>base_vseq.sv</code><br><code>isptx_sequence_1.sv:1104</code></td></tr>
      <tr><td>链路训练</td><td><code>send_isp()</code> 先发送 LANE_NUM transaction；1 lane 使用 TYPE_TRAINING，2 lane 使用 TYPE_TRAINING_TWO，并与 33 ms timeout 并行。</td><td><code>isptx_sequence_1.sv:1028</code></td></tr>
      <tr><td>逐帧组包</td><td>TRAIN_ACTIVE 调用 <code>send_ppm()</code>。每帧按 VBP setting/blank、VACT 的 HBP/BAC+POL/pixel/EOL/HFP、VFP 顺序创建 transaction；unlock monitor 与发送线程并行。</td><td><code>isptx_sequence_1.sv:423</code></td></tr>
      <tr><td>sequencer/driver</td><td>agent 将 driver <code>seq_item_port</code> 接到 sequencer export。driver 用 <code>get_next_item()</code> 取 transaction，按 id 分派 task，完成后调用 <code>item_done()</code>。</td><td><code>isptx_agent.sv</code><br><code>isptx_driver.sv</code></td></tr>
      <tr><td>bit-level 驱动</td><td>单 lane pixel 驱动 lane0，双 lane pixel 并行驱动 lane0/lane1。每个 9-bit symbol 在 9 个时钟上驱动差分 <code>rxpN/rxnN</code>；定向 case 可切到 <code>tx_clk_modify</code>。</td><td><code>convert_one_bit()</code><br><code>convert_sec_bit()</code></td></tr></table>
      <div class="note"><strong>transaction 不是整帧：</strong><code>isptx_transaction.id</code> 区分 LANE_NUM、TRAINING、REGISTER、HBK、BAC、POL、EOL、PIXEL 和异常 dummy。sequence 决定协议顺序与内容，driver 负责转换为引脚时序。</div>
      <h2>帧内协议与失锁恢复</h2><pre><code>TRAIN_SEND -> TYPE_TRAINING / TYPE_TRAINING_TWO -> wait LOCK
TRAIN_ACTIVE -> setting in VBP
             -> [HBP -> BAC+POL -> pixel -> EOL -> HFP] x VACT
             -> VFP
             || monitor PAD_LOCK_LEFT_IO == 0
unlock -> TRAIN_SEND; frames done -> DRV_PPM_DONE</code></pre>
      <h2>当前源码中的模型命令</h2>
      <pre><code>pic_process ... -W HACT -H VACT -P PORT_NUM -r subpix_reg \\
  -D DOTC -O pol -C POLC -S SHL -HV H120V -L DPLC_MODE

drdod_process.py ... --drd_panel DRD_PANEL --od_k OD_k \\
  --od_w1 ... --od_w6 ... --od_gray OD_gray

dplc.pl --input ... --mode ... --ave_r ... --ave_g ... \\
  --ave_b ... --delta ... --ave_last ...</code></pre>
      <span class="source">Source: top/agents/isprx_env/isptx_agent/isptx_sequence.sv</span>
      {figure('golden-flow.svg', 'Golden 数据流：普通、DPLC、DRDOD 模型生成文件，scoreboard 按配置选择并加载。')}
      <h2>CModel 生成 Golden 与 Checker 调用</h2>
      <ol class="steps"><li><strong>输入一致。</strong>sequence 使用当前 frame 的 <code>pattern/frame_N.ppm</code> 和同一份 rx_cfg 构造模型命令。</li><li><strong>普通 golden。</strong><code>pic_process</code> 输出到单芯片 <code>outResult</code> 或多芯片 <code>outResult1</code>，包含 Data Merge、Digital Top 和 Analog 层结果。</li><li><strong>特殊 golden。</strong>DPLC 生成 <code>DPLC_frameN_idM.ppm</code>；DRDOD 从 <code>input_ppm/drv_M_frame_N.ppm</code> 生成 <code>drdod_out_frameN.ppm</code>。</li><li><strong>checker 加载。</strong>scoreboard 不调用 C API，而是读取模型文件。Analog scoreboard 的 <code>get_cmodel()</code> 按 OD_k、DPLC_MODE 和普通模式选择输入。</li><li><strong>逐层比较。</strong>monitor transaction 形成 actual queue/PPM，scoreboard 按 frame/id、行和 subpixel 比较并报告 UVM error。</li></ol>
      <table><tr><th>模式</th><th>Golden 文件</th><th>Analog checker 行为</th></tr><tr><td>普通</td><td><code>outResult/analog/frame_N_analogM.txt</code><br>多芯片为 <code>outResultM/...</code></td><td>逐行解析空格分隔 subpixel，形成 model frame queue。</td></tr><tr><td>DPLC</td><td><code>DPLC_frameN_idM.ppm</code></td><td>读取 P3 PPM；关闭 channel 填 8'h22，SHL=0 时反转行。</td></tr><tr><td>DRDOD</td><td><code>drdod_out_frameN.ppm</code></td><td>读取 VACT*2 行；关闭 channel 填 8'h22，再与 analog DUT 输出比较。</td></tr></table>
      {figure('special-modes.svg', 'DPLC 与 DRDOD 使用不同的数据构造、模型文件和 checker 分支。')}
      <h2>DPLC 与 DRDOD 特殊处理</h2>
      <div class="grid-2"><section class="panel"><h3>DPLC</h3><ul><li><code>DPLC_MODE==1</code> 时运行 dplc.pl；当前源码把 mode 固定为 duplicate，并传 AVE_R/G/B、AVE_DELTA 和 AVE_LAST。</li><li>load_ppm 后调整 subpix_num；特定 subpix_reg=011、2 lane 路径会重排奇偶行和 pair 数据。</li><li>write_ppm 在 DPLC 模式交错写入两组 RGB，checker 改读 DPLC PPM。</li><li>Analog compare 对 subpix_reg=0 的 DPLC 分支存在 120 点偏移。</li></ul></section><section class="panel"><h3>DRDOD</h3><ul><li><code>OD_k != 0</code> 时由 <code>drdod_generate_pixel_data()</code> 按 DRD_PANEL、PAIR_NUM、CYCLIC/NOMATCH 直接构造 8-bit 数据并做 8B9B encode。</li><li>送入 DUT 的 decoded 数据写到 input_ppm，供 drdod_process.py 生成 golden。</li><li>Analog checker 行数由 VACT 变为 VACT*2，并读取 DRDOD PPM。</li><li>DRD input/output monitor 与 scoreboard 是独立检查链，验证 panel gate、输入模式和 output/bypass。</li></ul></section></div>
    """

    stimulus_body += f"""
      {figure('frame-sequence.svg', '逐帧时序：配置、模型、setting/pixel packet 与 scoreboard 必须共享同一个 frame 上下文。')}
      <h2>Case 资产清单</h2>
      <table><tr><th>资产</th><th>用途</th><th>常见错误</th></tr><tr><td><code>user_def.sv</code></td><td>静态尺寸、pair/port、帧数和功能宏</td><td>从其他 case 复制后未同步 HACT/VACT 或 CHIP_SEL</td></tr><tr><td><code>cfg_frame*.txt</code></td><td>每帧寄存器字段</td><td>文件数与 FRAME_NUM/复用规则不一致</td></tr><tr><td><code>pattern/frame_N.ppm</code></td><td>像素输入</td><td>宽高、P3 header、通道排列不匹配</td></tr><tr><td>test class</td><td>checker 开关、force、I2C 和异常注入</td><td>关闭 checker 后只看波形，没有自动判定</td></tr></table>
      <h2>Golden 产物核对</h2><ul><li>先从日志确认实际执行的命令行和参数，不凭文档猜测。</li><li>确认输出文件的修改时间属于本次 seed，防止读到旧产物。</li><li>确认 frame、pair、odd/even、left/right 命名与 scoreboard 加载路径一致。</li><li>模型报错或输出为空应直接使 case fail，不能继续比较空文件。</li></ul>
    """

    checkers_body = f"""
      {figure('checker-flow.svg', 'Checker 数据流：monitor analysis_port 连接 scoreboard，golden 来源依层级而异。')}
      <div class="note"><strong>组件创建方式：</strong><code>checker_agent.build_phase</code> 无条件创建 Data Merge、Digital Top、Chopper、Analog、DRD input 和 DRD output 的全部 monitor/scoreboard。各 <code>*_check_on</code> 在组件内部控制采样、文件读取或比较，不控制 factory create。</div>
      <div class="note"><strong>源码审计发现：</strong><code>digital_top_monitor.sv:31</code> 在 transaction 非空判断中使用 <code>env_cfg.data_merge_check_on</code>，而 scoreboard 比较使用 <code>digital_top_check_on</code>。关闭 Data Merge、只打开 Digital Top 时应确认该条件是否会影响预期采样。</div>
      {figure('checker-position.svg', 'Checker 在 DUT 数据通路中的 tap：主链逐层检查，DRD input/output 为并行独立路径。')}
      <h2>Checker 在 DUT 数据通路中的位置</h2>
      <table><tr><th>顺序</th><th>检查位置</th><th>观察对象</th><th>与其他组件的关系</th></tr><tr><td>1</td><td>Data Merge</td><td><code>data_merge_intf</code> 暴露的 merge 输出，两路数据流。</td><td>最靠近数字数据合并结果；它先失败通常说明输入解包、pair 合并或映射已错，后级失败可能是连锁结果。</td></tr><tr><td>2</td><td>Digital Top</td><td><code>digital_top_intf</code> 的 OL/EL/OR/ER 四路拆分输出。</td><td>位于 merge 后、chopper/analog 前；负责 odd/even 和 left/right 分流正确性。</td></tr><tr><td>3</td><td>Chopper</td><td><code>chopper_intf</code> 的 chopper dump 数据。</td><td>检查 digital 输出经过 chopper 处理后的中间结果；当前只连接 D port，G port 虽声明但未接 scoreboard。</td></tr><tr><td>4</td><td>Analog/control</td><td><code>analog_data_output_if</code> 的最终 pixel、POL、chopper、unlock、VBK。</td><td>最接近 source-driver 宏输出端；汇总最终图像与五类控制事件，并读取普通/DPLC/DRDOD golden。</td></tr><tr><td>并行支路</td><td>DRD input</td><td>DRD 模块输入侧 transaction。</td><td>验证 DRD_PANEL、OD 输入模式和 cyclic/nomatch pattern，帮助区分输入构造错误与 DRD 运算错误。</td></tr><tr><td>并行支路</td><td>DRD output</td><td>DRD 模块输出侧 transaction、DRDOD_EN/bypass。</td><td>与 DRD input 独立连接；bypass 场景比较 input/output，不能用 Analog checker 替代。</td></tr></table>
      <div class="note"><strong>定位原则：</strong>checker 不是串联调用关系，而是在 DUT 不同 tap 并行采样。调试时按数据通路找“第一个失败层”：Data Merge 正常而 Digital Top 失败，问题集中在拆分阶段；Digital Top 正常而 Analog 失败，优先查 chopper、channel mapping、控制时序或模型分支。</div>
      <h2>analysis_port 精确连接</h2>
      <table><tr><th>Monitor port</th><th>Scoreboard imp/export</th><th>数据流数量</th></tr>
      <tr><td><code>out_data_dump_aport</code> / <code>_1</code></td><td><code>out_data_dump_aexport</code> / <code>_1</code></td><td>Data Merge 2 路</td></tr>
      <tr><td><code>digital_top_ol/el/or/er_aport</code></td><td><code>out_data_ol/el/or/er_aexport</code></td><td>Digital Top 4 路</td></tr>
      <tr><td><code>chopper_dump_aport_d</code></td><td><code>out_data_chop_aexport_d</code></td><td>Chopper 1 路；<code>aport_g</code> 已声明但未在 checker_agent 连接</td></tr>
      <tr><td><code>mon2scb</code>、<code>pol</code>、<code>chopper</code>、<code>unlock</code>、<code>vbk</code></td><td><code>monitor_imp</code>、<code>pol_imp</code>、<code>chopper_imp</code>、<code>unlock_imp</code>、<code>vbk_imp</code></td><td>Analog/control 5 路</td></tr>
      <tr><td>DRD input/output 各自的 <code>out_aport</code></td><td>各自 scoreboard 的 <code>drdod_imp</code></td><td>DRD 2 条独立路径</td></tr></table>
      <h2>检查路径</h2>
      <table><tr><th>开关</th><th>采样/比较</th><th>关键事实</th></tr>
      <tr><td><code>data_merge_check_on</code></td><td>Data Merge monitor / scoreboard</td><td>两路 analysis port；输出 PPM，并读取 <code>outResult/data_merge</code> 文本。</td></tr>
      <tr><td><code>digital_top_check_on</code></td><td>Digital Top monitor / scoreboard</td><td>分别比较 OL、OR、EL、ER，错误含 frame/line/pixel 数据。</td></tr>
      <tr><td><code>chopper_check_on</code></td><td>Chopper monitor / scoreboard</td><td>检查 chopper dump 数据。</td></tr>
      <tr><td><code>analog_check_on</code></td><td>Analog output monitor / scoreboard</td><td>同时连接 pixel、POL、chopper、unlock、VBK 五类 analysis path。</td></tr>
      <tr><td><code>drd_input_check_on</code></td><td>DRD input scoreboard</td><td>根据 DRD_PANEL gate pattern、OD_k 和 cyclic mode 生成期望。</td></tr>
      <tr><td><code>drd_output_check_on</code></td><td>DRD output scoreboard</td><td>检查 DRDOD_EN，并在 bypass 时比较 input/output。</td></tr></table>
      <h2>失败定位顺序</h2>
      <ol class="steps"><li>先确认 testcase 中对应 <code>*_check_on</code> 是否真的打开。</li><li>检查 golden 文件是否存在，特别是 <code>outResult</code>、DPLC 和 DRDOD 输出。</li><li>按最前级失败定位：Data Merge -> Digital Top -> Chopper/Analog；DRDOD 使用独立 input/output 路径。</li><li>用日志中的 frame/line/pixel 定位波形，不要只看最终 PPM。</li></ol>
    """

    checkers_body += f"""
      {figure('debug-flow.svg', '失败定位决策：先验证 checker 和 golden，再寻找最早出错层级，最后回到单 case 复现。')}
      <h2>错误类型与优先证据</h2>
      <table><tr><th>现象</th><th>优先检查</th><th>保留产物</th></tr><tr><td>第一帧立即大量 mismatch</td><td>尺寸、通道排列、CHSEL/POLC/DOTC/SHL 与 golden 参数</td><td>命令行、cfg_frame0、actual/golden 首行</td></tr><tr><td>某一帧开始失败</td><td>frame index、cfg 复用、寄存器切换时刻</td><td>前后两帧 setting packet 与寄存器 dump</td></tr><tr><td>只有 OL/OR/EL/ER 一路失败</td><td>odd/even、left/right 拆分和文件选择</td><td>该路 monitor dump 与对应 golden</td></tr><tr><td>DRD bypass 失败</td><td>DRDOD_EN、panel gate、input/output 对齐</td><td>DRD input/output transaction</td></tr><tr><td>图像 checker 全关但 case fail</td><td>testcase 内 I2C、层次信号或 UVM error 检查</td><td>定向 task 日志与 force/release 时刻</td></tr></table>
    """

    run_body = f"""
      {figure('case-lifecycle.svg', 'Case 从模板到回归的推荐闭环。')}
      <h2>单 Case</h2><pre><code>cd $DV_TCON_C/script
run_tc.sh TOP &lt;case_name&gt; chip_tb_top random clean
run_tc.sh TOP &lt;case_name&gt; chip_tb_top verdi</code></pre>
      <p class="source">Source: script/run_tc.sh help and top/tests/run_case.py</p>
      <h2>回归</h2><pre><code>cd $DV_TCON_C/top/tests
python regression.py
# Coverage variant
python regression_cov.py</code></pre>
      <p><code>regression.py</code> 从 <code>case_list.txt</code> 读取 testcase，并生成 <code>bsub -Ip run_tc.sh TOP ... random clean</code> 命令。</p>
      <h2>提交回归前</h2><ul><li>case 名称与目录、主 .sv class、test_lib include 一致。</li><li>cfg_frame 数量覆盖 FRAME_NUM，或明确依赖复用规则。</li><li>checker 开关与验证目标一致，不能只依赖波形人工判断。</li><li>golden 命令在仿真目录生成预期文件，日志没有 file-open error。</li></ul>
    """

    run_body += """
      <h2>结果目录建议保留内容</h2><div class="grid-2"><section class="panel"><h3>复现必需</h3><ul><li>完整命令、seed、case 名和源码 revision。</li><li>compile/run log 与第一个 UVM error 上下文。</li><li>本次生成的 golden 和 actual dump。</li></ul></section><section class="panel"><h3>调试必需</h3><ul><li>波形数据库及保存范围。</li><li>cfg frame、setting packet、关键寄存器 dump。</li><li>checker 开关与模型命令行。</li></ul></section></div>
      <h2>回归清单治理</h2><p>当前 <code>case_list.txt</code> 的唯一名称与实际目录存在差异。新增清单项前应自动检查目录存在、主 class 可编译、输入文件齐全，并把不存在的名称单独报告，避免“提交了回归但实际没有运行目标”的假覆盖。</p>
    """

    case_types = sorted({str(item["caseType"]) for item in cases})
    feature1_values = sorted({str(item["feature1"]) for item in cases
                              if item["feature1"]})
    feature2_values = sorted({str(item["feature2"]) for item in cases
                              if item["feature2"]})
    case_type_options = "".join(
        f'<option value="{esc(value)}">{esc(value)}</option>'
        for value in case_types)
    feature1_options = "".join(
        f'<option value="{esc(value)}">{esc(value)}</option>'
        for value in feature1_values)
    feature2_options = "".join(
        f'<option value="{esc(value)}">{esc(value)}</option>'
        for value in feature2_values)
    hierarchy_rows = "".join(
        f"<tr><td>{esc(str(item['caseType']))}</td>"
        f"<td>{esc(str(item['feature1']) or 'Excel 未填写')}</td>"
        f"<td>{esc(str(item['feature2']))}</td>"
        f"<td>{item['namedCount']}</td><td><code>{esc(str(item['source']))}</code></td></tr>"
        for item in metadata["featureGroups"])
    cases_body = f"""
      <div class="stats"><div class="stat"><strong>{len(cases)}</strong><span>Excel plan rows</span></div><div class="stat"><strong>{passed_count}</strong><span>PASS rows</span></div><div class="stat"><strong>{linked_count}</strong><span>source linked</span></div><div class="stat"><strong id="case-count">0</strong><span>current results</span></div></div>
      <div class="note"><strong>数据来源：</strong>Case 范围与描述以 Excel 的 <code>normal case</code>、<code>waveform case</code> sheet 为准。空白描述/check 项按同一连续 Feature 分组继承上一条有效内容；每张卡保留 sheet 和行号便于回查。</div>
      <div class="note"><strong>源码关联：</strong>共 {linked_count} 条计划行能关联当前 tests 目录，{excel_only_count} 条仅存在于 Excel 或名称尚未与目录一致。源码关联不会覆盖 Excel 描述，只增加宏、checker、force 和检查调用证据。</div>
      <div class="case-controls"><input id="case-search" type="search" placeholder="搜索 case、Feature、描述、check 点、owner、宏或源码"><select id="case-type"><option value="">全部 Case Type</option>{case_type_options}</select><select id="case-feature1"><option value="">全部 Feature I</option><option value="__blank__">Feature I：Excel 未填写</option>{feature1_options}</select><select id="case-feature2"><option value="">全部 Feature II</option>{feature2_options}</select><select id="case-scope"><option value="all">全部计划项</option><option value="pass">仅 PASS</option><option value="linked">已关联源码</option><option value="missing">仅 Excel</option></select></div>
      <div class="case-results" id="case-results"></div>
      <h2>Normal / Waveform Feature 分组</h2><p>以下层级直接来自 Excel。Normal Case 同时列出 Feature I 与 Feature II；Waveform Case 的 Feature I 原表为空，因此如实标记为“Excel 未填写”，并完整列出六个 Feature II 分组。只有 Feature 标题但暂时没有 Case Name 的分组也会保留并显示为 0。</p><div class="table-scroll"><table><tr><th>Case Type</th><th>Feature I</th><th>Feature II</th><th>命名 Case 数量</th><th>Excel 位置</th></tr>{hierarchy_rows}</table></div>
      <h2>代表性 Case 阅读方法</h2>
      <div class="grid-2"><section class="panel"><h3>基础数据通路</h3><p>从 <code>t_8b1lane</code> 开始，核对 PAIR_NUM、PORT_NUM、尺寸宏和四类基础图像 checker，再与 2-lane/多 port 变体比较。</p></section><section class="panel"><h3>DRDOD</h3><p>查看 <code>t_8b2lane_DRD_PANEL3</code> 与 <code>t_drdod_en_toggle_2lane_HKC1_R</code>，重点跟踪 DRD 功能宏、panel 配置和 input/output checker。</p></section><section class="panel"><h3>协议异常</h3><p>查看 <code>t_pixel_with_training_pattern</code>、prefix/setting 异常类 case，结合 force、注入参数和关闭的 checker 理解定向判定。</p></section><section class="panel"><h3>寄存器 / I2C</h3><p>查看 <code>t_i2c_access_reg_unlock_reset</code>，从 I2C_SIM、check 调用和 force/release 证据理解非图像类验证。</p></section></div>
      <script src="assets/cases-data.js?v=20260727"></script><script src="assets/guide.js?v=20260727"></script>
    """

    register_rows = "".join(
        "<tr>" + "".join(
            f"<td>{cell_html(item[key])}</td>"
            for key in ("address", "name", "type", "default", "access",
                        "description", "values", "simCheck"))
        + "<td>" + (
            f'<details><summary>{len(item["caseNames"])} 个 case</summary>'
            '<div class="register-case-links">'
            + "".join(
                f'<a href="cases.html?q={quote(name)}"><code>{esc(name)}</code></a>'
                for name in item["caseNames"])
            + "</div></details>"
            if item["caseNames"] else '<span class="muted">未关联</span>'
        ) + "</td><td>" + cell_html(item["linkMethod"]) + "</td></tr>"
        for item in metadata["registers"])
    change_rows = "".join(
        f"<tr><td>{cell_html(item['date'])}</td><td>{cell_html(item['file'])}</td>"
        f"<td>{cell_html(item['action'])}</td><td>{cell_html(item['owner'])}</td>"
        f"<td>{cell_html(item['status'])}</td><td>{cell_html(item['note'])}</td></tr>"
        for item in metadata["changes"])
    format_rows = "".join(
        f"<tr><td>{cell_html(item['no'])}</td><td>{cell_html(item['hPeriod'])}x{cell_html(item['vPeriod'])}</td>"
        f"<td>{cell_html(item['hBlank'])}/{cell_html(item['vBlank'])}</td><td>{cell_html(item['fps'])}</td>"
        f"<td>{cell_html(item['depth'])}</td><td>{cell_html(item['channel'])}</td><td>{cell_html(item['pcs'])}</td>"
        f"<td>{cell_html(item['pair'])}</td><td>{cell_html(item['ispSpeed'])}</td><td>{cell_html(item['mode'])}</td></tr>"
        for item in metadata["formats"])
    status_rows = "".join(
        f"<tr><td>{cell_html(item['type'])}</td><td>{cell_html(item['pass'])}</td>"
        f"<td>{cell_html(item['error'])}</td><td>{cell_html(item['ongoing'])}</td>"
        f"<td>{cell_html(item['unbuilt'])}</td><td>{cell_html(item['total'])}</td></tr>"
        for item in metadata["caseStatus"])
    history_rows = "".join(
        f"<tr><td>{cell_html(item['date'])}</td><td>{cell_html(item['comment'])}</td>"
        f"<td>{cell_html(item['owner'])}</td></tr>"
        for item in metadata["history"])
    coverage = metadata["coverage"]
    linked_registers = sum(bool(item["caseNames"])
                           for item in metadata["registers"])
    no_case_registers = sum(
        str(item["simCheck"]).upper() == "NO CASE"
        for item in metadata["registers"])
    pending_registers = (len(metadata["registers"]) - linked_registers
                         - no_case_registers)
    env_status_counts = Counter(
        str(item["status"]).lower() or "未填写"
        for item in metadata["changes"])
    env_owner_counts = Counter(
        str(item["owner"]) or "未填写" for item in metadata["changes"])
    env_status_text = " / ".join(
        f"{name}: {count}" for name, count in env_status_counts.most_common())
    env_owner_text = " / ".join(
        f"{name}: {count}" for name, count in env_owner_counts.most_common())
    coverage_metrics = [
        ("Line", coverage["line"]), ("Condition", coverage["condition"]),
        ("Toggle", coverage["toggle"]), ("FSM", coverage["fsm"]),
        ("Branch", coverage["branch"]),
    ]
    weakest_coverage = min(
        coverage_metrics,
        key=lambda item: float(item[1]) if item[1] else float("inf"))
    plan_body = f"""
      <div class="note"><strong>唯一主清单：</strong><code>{cell_html(metadata['workbook'])}</code>。本页保留 workbook 中除 case 明细外的 register、ENV changelist、coverage 和 video format 信息；case 明细见 <a href="cases.html">Case 计划索引</a>。</div>
      <h2>计划状态</h2><div class="stats"><div class="stat"><strong>{len(cases)}</strong><span>named rows</span></div><div class="stat"><strong>{passed_count}</strong><span>PASS</span></div><div class="stat"><strong>{len(metadata['registers'])}</strong><span>register fields</span></div><div class="stat"><strong>{len(metadata['formats'])}</strong><span>video formats</span></div></div>
      <p>下表直接保留 workbook 的 <code>case status</code> 统计口径。命名计划行还包含未计入该状态表的标题或扩展条目，因此网页索引总数与状态表 Total 不要求相等。</p><table><tr><th>Case type</th><th>Pass</th><th>Error</th><th>On going</th><th>Unbuild</th><th>Total</th></tr>{status_rows}</table>
      <h2>Workbook 维护历史</h2><table><tr><th>Date</th><th>Comment</th><th>Owner</th></tr>{history_rows}</table>
      <h2>Coverage 快照</h2><table><tr><th>Score</th><th>Line</th><th>Condition</th><th>Toggle</th><th>FSM</th><th>Branch</th><th>Date</th></tr><tr><td>{cell_html(coverage['score'])}%</td><td>{cell_html(coverage['line'])}%</td><td>{cell_html(coverage['condition'])}%</td><td>{cell_html(coverage['toggle'])}%</td><td>{cell_html(coverage['fsm'])}%</td><td>{cell_html(coverage['branch'])}%</td><td>{cell_html(coverage['date'])}</td></tr></table><div class="note"><strong>Closure 重点：</strong>当前最低项为 {esc(weakest_coverage[0])} {cell_html(weakest_coverage[1])}%。该表是 Excel 在 {cell_html(coverage['date'])} 记录的单次快照，不代表实时回归结果；下一次更新应同时记录回归版本、waiver 和未覆盖原因。</div>
      <h2>Video Format 矩阵</h2><p>用于核对分辨率、blank、帧率、色深、driver channel、PCS、pair 和单 pair 速率。Case 名中的简化参数不能替代本表的系统带宽条件。</p><div class="table-scroll"><table><tr><th>No</th><th>Active</th><th>H/V blank</th><th>FPS</th><th>Depth</th><th>Channel</th><th>PCS</th><th>Pair/Driver</th><th>iSP Speed/Pair</th><th>Mode</th></tr>{format_rows}</table></div>
      <h2>寄存器验证映射</h2><div class="stats"><div class="stat"><strong>{len(metadata['registers'])}</strong><span>register fields</span></div><div class="stat"><strong>{linked_registers}</strong><span>linked to cases</span></div><div class="stat"><strong>{no_case_registers}</strong><span>NO CASE</span></div><div class="stat"><strong>{pending_registers}</strong><span>manual / unresolved</span></div></div><p>关联依据保留在“关联规则”列：优先使用 Excel 中的精确 testcase 名和通配模式，其次使用 SHL、CHSEL、DOTC、POLC、H120V、DRDOD、DPLC、UTC 等明确功能关键词。无法可靠推断的条目保持未关联。点击 case 名可跳转并自动筛选 Case 计划索引。</p><details><summary>展开全部 {len(metadata['registers'])} 个寄存器字段</summary><div class="table-scroll"><table class="register-table"><tr><th>Address</th><th>Name</th><th>Type</th><th>Default</th><th>Access</th><th>Description</th><th>Values</th><th>Sim Check</th><th>关联 Case</th><th>关联规则</th></tr>{register_rows}</table></div></details>
      <h2>环境修改记录</h2><div class="grid-2"><section class="panel"><h3>状态分布</h3><p>{esc(env_status_text)}</p></section><section class="panel"><h3>Owner 分布</h3><p>{esc(env_owner_text)}</p></section></div><p>这些记录反映环境能力在哪个版本引入，也是移植下一项目时优先审计的风险列表。优先检查状态未完成、文件名为空或描述依赖旧层次路径的记录。</p><details open><summary>展开 {len(metadata['changes'])} 条环境变更</summary><div class="table-scroll"><table><tr><th>Date</th><th>File</th><th>Action</th><th>Owner</th><th>Status</th><th>Note</th></tr>{change_rows}</table></div></details>
    """

    portability_body = """
      <h2>复用目标</h2><p>下一项目不应复制整个目录后逐个修编译错误，而应把环境拆成稳定平台层、芯片适配层和项目验证内容层。平台层保持协议与 UVM 机制稳定，所有 RTL 层次、寄存器表、分辨率和模型差异集中到适配层。</p>
      <div class="grid-3"><section class="panel"><h3>平台层</h3><p>ISPTX/I2C agent、transaction、sequencer、通用 packet 发送、analysis port、回归框架和报告格式。</p></section><section class="panel"><h3>项目适配层</h3><p>TB top、interface bind、DUT hierarchy tap、env_cfg、寄存器打包、golden 命令和 checker 数据映射。</p></section><section class="panel"><h3>验证内容层</h3><p>Excel 计划、testcase、cfg frame、pattern、功能 checkpoint、waiver 和 coverage closure。</p></section></div>
      <h2>当前环境中的耦合点</h2><table><tr><th>耦合点</th><th>当前位置</th><th>移植策略</th></tr><tr><td>DUT 层次与 checker tap</td><td><code>chip_tb_top.sv</code>、interface/connect、定向 force</td><td>集中到 bind/interface adapter，禁止通用 monitor 散落新项目层次路径。</td></tr><tr><td>连接数量</td><td><code>CONNECT_NUM</code> 分支分散在 TB、base_test、base_vseq</td><td>改为数组/循环或统一 connection descriptor；修正 2/3 路额外 sequence 风险。</td></tr><tr><td>寄存器配置</td><td><code>env_cfg.sv</code>、<code>process_cfg()</code>、<code>get_reg()</code></td><td>从机器可读 register spec 生成字段模型与 packer，保留项目覆盖层。</td></tr><tr><td>图像与 golden</td><td><code>isptx_sequence*.sv</code> 内拼接命令</td><td>抽象 model adapter API，参数对象化；路径和输出命名由项目配置提供。</td></tr><tr><td>Checker 文件路径</td><td>scoreboard 内部 <code>outResult</code> 和 PPM 命名</td><td>统一 artifact locator，输入 frame/id/path，避免 scoreboard 自行拼字符串。</td></tr><tr><td>Case 计划</td><td>Excel 与 tests 目录</td><td>Excel 保持需求主清单；生成稳定 case ID，与源码和 coverage 双向关联。</td></tr></table>
      <h2>移植到下一项目的八个阶段</h2><ol class="steps"><li><strong>冻结基线。</strong>记录当前回归 PASS 数、coverage、模型版本、寄存器版本和已知问题。</li><li><strong>建立适配清单。</strong>列出接口、时钟复位、连接数、lane/pair/port、寄存器、模型和 DUT tap 差异。</li><li><strong>先移植 TB 静态层。</strong>完成 DUT 实例、clock/reset、ISPTX/I2C interface 和 checker tap，执行 interface smoke。</li><li><strong>移植配置层。</strong>生成新 env_cfg/register packer，完成 default、RW、unlock/reset 和多帧配置测试。</li><li><strong>打通单连接主链路。</strong>只启用 Data Merge，验证 setting、pixel、golden、actual 和 UVM error 闭环。</li><li><strong>逐层打开 checker。</strong>Digital Top -> Chopper -> Analog -> DRD input/output，每层先做正向 PASS 和故意 mismatch。</li><li><strong>扩展多连接和格式矩阵。</strong>验证 connection id、文件隔离、并行 sequence 和带宽边界，不假设旧 CONNECT_NUM 分支正确。</li><li><strong>迁移验证计划。</strong>逐行映射 Excel Feature/checkpoint 到新项目 case ID，标记复用、修改、新增或不适用，并重新关闭 coverage。</li></ol>
      <h2>移植验收门槛</h2><table><tr><th>阶段</th><th>必须通过</th><th>禁止项</th></tr><tr><td>Compile/Elaboration</td><td>无层次路径、interface、宏和 config_db fatal</td><td>不得用 force 绕过结构连接问题</td></tr><tr><td>Agent smoke</td><td>ISPTX/I2C transaction 数和时序符合预期</td><td>不得关闭 monitor 掩盖流量缺失</td></tr><tr><td>Golden closure</td><td>模型命令可复现，产物按 frame/id 隔离</td><td>不得读取旧项目残留文件</td></tr><tr><td>Checker closure</td><td>每条 analysis path 有正向与负向实验</td><td>只 PASS、不能故意 fail 不算完成</td></tr><tr><td>Regression</td><td>Excel 计划行有明确映射和自动结果</td><td>Excel-only、source-only 和 waiver 必须有 owner</td></tr><tr><td>Coverage</td><td>达到新项目目标并解释差异</td><td>不能直接继承旧项目 coverage 数字</td></tr></table>
      <h2>建议的目录边界</h2><pre><code>dv_platform/        # reusable agents, transactions, base sequences
project_adapter/    # DUT interfaces, register/model/checker adapters
project_tests/      # plan IDs, tests, cfg frames, patterns
project_config/     # paths, connection descriptors, format matrix
artifacts/          # run-id/frame-id scoped golden and actual outputs</code></pre>
      <div class="note"><strong>完成定义：</strong>移植完成不是“smoke case 能跑”，而是 Excel 计划可追溯、每条 checker 能主动抓错、多连接和关键 video format 通过、coverage 重新关闭，并且没有依赖旧项目绝对路径或残留产物。</div>
    """

    faq_body = """
      <h2>为什么 case 名写 CHSEL1，但 CHIP_SEL 是 0？</h2><p>CHSEL 是数据映射寄存器参数；CHIP_SEL 是选择芯片寄存器映射版本的编译宏。两者独立，必须分别查看目录名/cfg_frame 与 user_def.sv。</p>
      <h2>为什么有些 case 没有图像 checker？</h2><p>定向 reset、WAKE、I2C 或异常协议 case 可能主动关闭图像 checker，转而在 testcase 中调用 check_i2c、check_cfg_signal 或直接检查层次信号。Case 索引会显示开关和源码调用计数。</p>
      <h2>为什么 cfg_frame 数量少于 FRAME_NUM？</h2><p>部分基础 case 只有 cfg_frame0；env_cfg/sequence 可能复用配置。新增 case 时不要假定复用，应先核对 process_cfg() 和实际仿真日志。</p>
      <h2>如何判断文档是否过期？</h2><p>页面头部显示审计日期；运行 <code>python tools/build_guide.py</code> 会重新读取 Excel 主清单并扫描 tests 源码。Excel 更新决定 case 范围和计划内容，env/checker 源码更新决定实现证据，两边都需要复核。</p>
      <h2>新人最短路径</h2><ol class="steps"><li>复制最接近的 testcase，而不是从空目录开始。</li><li>先让单 case deterministic PASS。</li><li>确认目标 checker 打开且能故意制造一次 fail。</li><li>再加入 case_list.txt 跑回归。</li></ol>
      <h2>CONNECT_NUM、PAIR_NUM 和 PORT_NUM 怎么区分？</h2><p><code>CONNECT_NUM</code> 决定创建多少套 source env、绑定多少组 interface，并选择多少路 sequence；<code>PAIR_NUM</code> 和 <code>PORT_NUM</code> 参与频率计算、图像宽度、子像素及 packet 数据组织。三个宏可能取相近数值，但职责不同，不能互相替代。</p>
      <h2>为什么要故意制造一次 fail？</h2><p>一个始终 PASS 的 case 不能证明 checker 真正连接且覆盖了目标路径。最小负向实验可以确认 monitor 有采样、scoreboard 加载了正确 golden，并且 mismatch 能传播成 UVM error。</p>
      <h2>回归通过后哪些文件值得保留？</h2><p>至少保留命令、seed、revision、compile/run log、checker 摘要和失败 case 的 golden/actual；定向协议或寄存器 case 还应保存 force/I2C 时序证据。</p>
    """

    return {
        "index.html": page("index.html", "验证环境使用指南", "以当前 DV_TCON_C 源码为事实源，快速理解环境、运行 testcase、定位 checker 失败。", index_body),
        "overview.html": page("overview.html", "环境概览与事实源", "说明配置、激励、golden 与检查结果如何形成闭环，并纠正旧文档中的概念混用。", overview_body),
        "tb-arch.html": page("tb-arch.html", "Testbench 架构", "依据 base_test、source_driver_env、base_vseq 和 checker_agent 的实际构建与连接关系。", arch_body),
        "stimulus.html": page("stimulus.html", "激励、配置与 Golden", "从 testcase 资产到寄存器包、pixel stream 和三类模型输出的逐帧执行路径。", stimulus_body),
        "checkers.html": page("checkers.html", "Checker 与失败定位", "按 Data Merge、Digital Top、Chopper/Analog 和 DRDOD 独立路径说明实际 monitor/scoreboard。", checkers_body),
        "plan.html": page("plan.html", "Excel 验证计划总览", "整理主清单中的计划状态、寄存器验证映射、环境变更、coverage 和 video format。", plan_body),
        "run.html": page("run.html", "运行、回归与 Case 生命周期", "使用仓库中的 run_tc.sh、regression.py 和 case_list.txt，建立可复现的运行闭环。", run_body),
        "cases.html": page("cases.html", "Testcase 验证计划索引", "以 HV2M23 EDA Excel 为主清单，整合验证目标、check 点、状态与可关联的源码证据。", cases_body),
        "portability.html": page("portability.html", "环境复用与下一项目移植", "把稳定平台机制与芯片适配内容分层，给出可执行的移植阶段、风险点和验收门槛。", portability_body),
        "faq.html": page("faq.html", "FAQ 与新人检查表", "集中解释 CHSEL/CHIP_SEL、checker 开关、配置帧复用和文档刷新方式。", faq_body),
    }


def main() -> None:
    """Generate assets, testcase data, and all HTML pages."""
    if not TESTS_DIR.exists():
        raise FileNotFoundError(f"Test directory not found: {TESTS_DIR}")
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    source_cases = scan_source_cases()
    cases, metadata = load_excel_plan(source_cases)
    category_counts = Counter(str(item["category"]) for item in cases)
    (ASSET_DIR / "guide.css").write_text(
        "/* Auto-generated by tools/build_guide.py */\n" + CSS.strip() + "\n",
        encoding="utf-8")
    (ASSET_DIR / "guide.js").write_text(
        "// Auto-generated by tools/build_guide.py\n" + GUIDE_JS.strip() + "\n",
        encoding="utf-8")
    data = json.dumps(cases, ensure_ascii=False, separators=(",", ":"))
    (ASSET_DIR / "cases-data.js").write_text(
        f"// Auto-generated by tools/build_guide.py\nwindow.HV2_CASES={data};\n",
        encoding="utf-8")
    build_svgs()
    for filename, content in build_pages(cases, metadata).items():
        (GUIDE_DIR / filename).write_text(content, encoding="utf-8")
    print(
        f"Generated {len(cases)} testcase records across "
        f"{len(category_counts)} categories."
    )


if __name__ == "__main__":
    main()
