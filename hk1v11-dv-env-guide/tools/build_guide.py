#!/usr/bin/env python3
"""Build the HK1V11 DV guide from the case plan and verification sources."""

from __future__ import annotations

import html
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


GUIDE_DIR = Path(__file__).resolve().parents[1]
ASSET_DIR = GUIDE_DIR / "assets"
LEGACY_DIR = ASSET_DIR / "legacy"
LEGACY_CONTENT_DIR = Path(__file__).resolve().parent / "legacy-content"
SOURCE_CANDIDATES = (
    Path(r"E:\HK1V11_new\DV_TCON_C"),
    Path(r"E:\HK1V11_github\DV_TCON_C"),
    Path(r"E:\HK1V11\DV_TCON_C"),
    Path(r"C:\Users\xiapeng2\Desktop\HK1V11\DV_TCON_C"),
)
CASELIST_XLSX = Path(
    r"C:\Users\xiapeng2\Desktop\HK1V11\04.Architecture\IP_Digital"
    r"\2.Verification\1.EDA\3.HK1V11_EDA_case_list.xlsx"
)
BUILD_DATE = date.today().isoformat()

NAV_ITEMS = [
    ("index.html", "首页"),
    ("overview.html", "概览"),
    ("tb-arch.html", "TB 架构"),
    ("stimulus.html", "激励与 Golden"),
    ("checkers.html", "检查机制"),
    ("plan.html", "验证计划"),
    ("run.html", "运行与回归"),
    ("scripts.html", "脚本指南"),
    ("cases.html", "Testcase 索引"),
    ("faq.html", "FAQ"),
]

LEGACY_PAGES = {
    "index.html": ("HK1V11 DV 环境指南", "原有环境说明与新增计划数据的统一入口。"),
    "overview.html": ("环境概览", "验证目标、目录结构、职责边界和端到端数据流。"),
    "tb-arch.html": ("TB 架构", "test、environment、agent、interface 与 DUT 的层次和连接关系。"),
    "stimulus.html": ("激励与 Golden", "ISPTX 激励、寄存器配置、pattern 和 CModel golden 数据路径。"),
    "checkers.html": ("检查机制", "各级 monitor、scoreboard、checksum 和数据比较链路。"),
    "run.html": ("运行与回归", "单 case、批量回归、日志和 coverage 工作流。"),
    "faq.html": ("FAQ", "环境使用限制、常见问题和执行前检查清单。"),
}

MOJIBAKE_MARKERS = "ÃÂæçåðñòóôõöøùúûüýþ鐢ㄤ緥姒傝妫€"


def find_source_root() -> Path:
    """Select the available HK1V11 source tree with the most testcase directories."""
    available: list[tuple[int, Path]] = []
    for root in SOURCE_CANDIDATES:
        tests_dir = root / "top" / "tests"
        if tests_dir.exists():
            case_count = sum(path.is_dir() for path in tests_dir.glob("t_*"))
            available.append((case_count, root))
    if not available:
        return SOURCE_CANDIDATES[0]
    return max(available, key=lambda item: item[0])[1]


SOURCE_ROOT = find_source_root()
TESTS_DIR = SOURCE_ROOT / "top" / "tests"


def esc(value: object) -> str:
    """Escape a value for HTML output."""
    return html.escape(str(value), quote=True)


def clean_text(value: object) -> str:
    """Normalize spreadsheet values and repair common legacy encodings."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\r\n", "\n").strip()
    candidates = [text]
    for encoding, decoding in (("latin1", "gb18030"), ("gb18030", "utf-8")):
        try:
            candidates.append(text.encode(encoding).decode(decoding))
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass

    def score(candidate: str) -> tuple[int, int]:
        marker_count = sum(candidate.count(char) for char in MOJIBAKE_MARKERS)
        replacement_count = candidate.count("�") + candidate.count("?")
        return marker_count + replacement_count * 4, -sum(
            "\u4e00" <= char <= "\u9fff" for char in candidate
        )

    return min(candidates, key=score)


def read_text(path: Path) -> str:
    """Read source text while tolerating legacy encodings."""
    for encoding in ("utf-8", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return ""


def write_text(path: Path, content: str) -> None:
    """Write generated UTF-8 text with a standard header where applicable."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8", newline="\n")


def archive_legacy_svgs() -> list[dict[str, str]]:
    """Preserve every inline SVG from the pre-generated guide."""
    LEGACY_DIR.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, str]] = []
    for page in sorted(GUIDE_DIR.glob("*.html")):
        if page.name in {"plan.html", "scripts.html"}:
            continue
        source = read_text(page)
        for index, match in enumerate(
            re.finditer(r"<svg\b[\s\S]*?</svg>", source, flags=re.IGNORECASE), start=1
        ):
            name = f"{page.stem}-{index:02d}.svg"
            target = LEGACY_DIR / name
            if not target.exists():
                write_text(target, match.group(0))
            manifest.append({"page": page.name, "asset": f"assets/legacy/{name}"})
    colors = {
        "var(--bg)": "#f4f6f8",
        "var(--surface)": "#ffffff",
        "var(--fg)": "#1b232b",
        "var(--text)": "#1b232b",
        "var(--muted)": "#66727d",
        "var(--accent)": "#087e6b",
        "var(--accent-soft)": "#dff3ee",
        "var(--border)": "#d9dee3",
        "var(--line)": "#d9dee3",
        "var(--pass)": "#16805f",
        "var(--pass-soft)": "#e0f3eb",
        "var(--warn)": "#a85d00",
        "var(--warn-soft)": "#fff0d8",
    }
    for target in LEGACY_DIR.glob("*.svg"):
        source = read_text(target)
        for token, color in colors.items():
            source = source.replace(token, color)
        write_text(target, source)
    archived_assets = {item["asset"] for item in manifest}
    for target in sorted(LEGACY_DIR.glob("*.svg")):
        asset = f"assets/legacy/{target.name}"
        if asset not in archived_assets:
            manifest.append({"page": target.stem.rsplit("-", 1)[0] + ".html", "asset": asset})
    write_text(
        LEGACY_DIR / "manifest.json",
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
    )
    return manifest


def archive_legacy_content() -> None:
    """Preserve original page bodies as stable inputs for the unified UI."""
    LEGACY_CONTENT_DIR.mkdir(parents=True, exist_ok=True)
    for page_name in LEGACY_PAGES:
        target = LEGACY_CONTENT_DIR / page_name
        if target.exists():
            continue
        source = read_text(GUIDE_DIR / page_name)
        if page_name == "index.html":
            match = re.search(
                r'<main class="content">([\s\S]*?)</main>', source, re.IGNORECASE
            )
        else:
            match = re.search(
                r'<article class="article">([\s\S]*?)</article>',
                source,
                re.IGNORECASE,
            )
        if not match:
            raise ValueError(f"Unable to preserve legacy content from {page_name}")
        write_text(
            target,
            "<!-- Preserved by tools/build_guide.py. -->\n" + match.group(1).strip() + "\n",
        )


def build_legacy_pages() -> None:
    """Render preserved guide content with the shared UI and navigation."""
    for page_name, (title, lead) in LEGACY_PAGES.items():
        content = read_text(LEGACY_CONTENT_DIR / page_name)
        if page_name == "tb-arch.html":
            content = """
<section class="source-upgrade">
  <h2>源码级 TB 层次与连接</h2>
  <p>依据 <code>base_test.sv</code>、<code>base_vseq.sv</code> 和 <code>env.sv</code> 补充。MULTI_CHIP 下由 CONNECT_NUM 控制 2~4 个 source_driver_env。</p>
  <figure class="diagram"><img src="assets/hk1v11-architecture.svg" alt="HK1V11 source-aligned UVM architecture"></figure>
</section>
""" + content
        elif page_name == "checkers.html":
            content = """
<section class="source-upgrade">
  <h2>源码级 Checker 连接</h2>
  <p>依据 <code>checker_agent.sv</code> 的 build/connect phase，区分 Data Merge、Digital Top、Chopper 和 Analog 四条检查链。</p>
  <figure class="diagram"><img src="assets/hk1v11-checker-architecture.svg" alt="HK1V11 source-aligned checker architecture"></figure>
</section>
""" + content
        body = f'<div class="legacy-content">{content}</div>'
        write_text(GUIDE_DIR / page_name, page_shell(title, page_name, body, lead))


def svg_document(width: int, height: int, body: str, title: str) -> str:
    """Wrap source-aligned diagram content in a standalone SVG document."""
    return f"""<!-- Auto-generated by tools/build_guide.py. -->
<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-label="{esc(title)}">
<defs><marker id="arrow" markerWidth="9" markerHeight="9" refX="8" refY="4.5" orient="auto"><path d="M0 0L9 4.5L0 9Z" fill="#65727d"/></marker></defs>
<style>.box{{fill:#fff;stroke:#70808d;stroke-width:1.4}}.green{{fill:#dff3ee;stroke:#087e6b;stroke-width:1.6}}.blue{{fill:#e5eef7;stroke:#245c91;stroke-width:1.5}}.amber{{fill:#fff0d8;stroke:#a85d00;stroke-width:1.5}}.group{{fill:#f4f6f8;stroke:#9aa7b1;stroke-width:1.2;stroke-dasharray:6 4}}.line{{fill:none;stroke:#65727d;stroke-width:1.5;marker-end:url(#arrow)}}.dash{{fill:none;stroke:#087e6b;stroke-width:1.5;stroke-dasharray:5 4;marker-end:url(#arrow)}}text{{font-family:Segoe UI,Microsoft YaHei,sans-serif;fill:#1b232b;font-size:15px}}.small{{font-size:12px;fill:#5f6d78}}.title{{font-size:17px;font-weight:700}}</style>
{body}
</svg>
"""


def build_source_diagrams() -> None:
    """Write HK1V11 architecture and checker diagrams from audited source facts."""
    architecture = """
<text class="title" x="36" y="34">HK1V11 UVM verification hierarchy</text>
<rect class="blue" x="35" y="65" width="190" height="78" rx="6"/><text x="130" y="94" text-anchor="middle">testcase / base_test</text><text class="small" x="130" y="118" text-anchor="middle">build env, reg model, coverage</text>
<rect class="blue" x="270" y="65" width="180" height="78" rx="6"/><text x="360" y="94" text-anchor="middle">sd_vsqr / base_vseq</text><text class="small" x="360" y="118" text-anchor="middle">ISPTX + I2C handles</text>
<path class="line" d="M225 104H270"/>
<rect class="group" x="500" y="48" width="660" height="390" rx="9"/><text class="title" x="520" y="78">source_driver_env [0..CONNECT_NUM-1]</text><text class="small" x="520" y="100">sd_env always; sd_env1..3 under MULTI_CHIP</text>
<path class="line" d="M450 104H500"/>
<rect class="green" x="535" y="130" width="170" height="72" rx="6"/><text x="620" y="159" text-anchor="middle">isptx_agent</text><text class="small" x="620" y="181" text-anchor="middle">pixel / setting stimulus</text>
<rect class="green" x="740" y="130" width="170" height="72" rx="6"/><text x="825" y="159" text-anchor="middle">i2c_agent</text><text class="small" x="825" y="181" text-anchor="middle">register transactions</text>
<rect class="green" x="945" y="130" width="170" height="72" rx="6"/><text x="1030" y="159" text-anchor="middle">bcc_agent</text><text class="small" x="1030" y="181" text-anchor="middle">reg model default map</text>
<rect class="amber" x="535" y="250" width="265" height="112" rx="6"/><text x="667" y="279" text-anchor="middle">checker_agent</text><text class="small" x="667" y="303" text-anchor="middle">Data Merge 1 / Digital Top 4</text><text class="small" x="667" y="325" text-anchor="middle">Chopper 1 / Analog 4 ports</text><text class="small" x="667" y="347" text-anchor="middle">monitor -> scoreboard</text>
<rect class="box" x="845" y="250" width="270" height="112" rx="6"/><text x="980" y="279" text-anchor="middle">virtual interfaces</text><text class="small" x="980" y="303" text-anchor="middle">data_merge_intf_0</text><text class="small" x="980" y="325" text-anchor="middle">digital_top_intf_0</text><text class="small" x="980" y="347" text-anchor="middle">chopper_intf_0</text>
<path class="dash" d="M845 306H800"/>
<rect class="blue" x="330" y="505" width="260" height="100" rx="6"/><text x="460" y="535" text-anchor="middle">chip_tb_top / HK1V11 DUT</text><text class="small" x="460" y="559" text-anchor="middle">ISPTX pads + I2C + BCC</text><text class="small" x="460" y="581" text-anchor="middle">digital / analog observation taps</text>
<path class="line" d="M620 202V470H510V505"/><path class="line" d="M825 202V455H480V505"/><path class="line" d="M1030 202V440H450V505"/>
<rect class="amber" x="690" y="505" width="300" height="100" rx="6"/><text x="840" y="535" text-anchor="middle">report / artifacts</text><text class="small" x="840" y="559" text-anchor="middle">outResult + input_ppm + reg dump</text><text class="small" x="840" y="581" text-anchor="middle">UVM_ERROR -> PASSED / FAILED</text>
<path class="line" d="M590 555H690"/>
<text class="small" x="36" y="675">Source: top/tb/base_test.sv; top/vseq/base_vseq.sv; top/agents/isprx_env/env.sv</text>
"""
    checkers = """
<text class="title" x="36" y="34">HK1V11 checker_agent build/connect topology</text>
<rect class="group" x="30" y="55" width="1140" height="585" rx="9"/><text class="title" x="50" y="86">checker_agent (created unconditionally)</text>
<rect class="blue" x="65" y="125" width="200" height="80" rx="6"/><text x="165" y="155" text-anchor="middle">Data Merge tap</text><text class="small" x="165" y="180" text-anchor="middle">data_merge_intf_0</text>
<rect class="green" x="330" y="125" width="245" height="80" rx="6"/><text x="452" y="155" text-anchor="middle">data_merge_monitor</text><text class="small" x="452" y="180" text-anchor="middle">out_data_dump_aport</text>
<rect class="amber" x="650" y="125" width="250" height="80" rx="6"/><text x="775" y="155" text-anchor="middle">data_merge_scoreboard</text><text class="small" x="775" y="180" text-anchor="middle">out_data_dump_aexport</text><path class="line" d="M265 165H330"/><path class="line" d="M575 165H650"/>
<rect class="blue" x="65" y="245" width="200" height="92" rx="6"/><text x="165" y="275" text-anchor="middle">Digital Top taps</text><text class="small" x="165" y="299" text-anchor="middle">OL / EL / ER / OR</text><text class="small" x="165" y="320" text-anchor="middle">digital_top_intf_0</text>
<rect class="green" x="330" y="245" width="245" height="92" rx="6"/><text x="452" y="275" text-anchor="middle">digital_top_monitor</text><text class="small" x="452" y="299" text-anchor="middle">4 analysis ports</text>
<rect class="amber" x="650" y="245" width="250" height="92" rx="6"/><text x="775" y="275" text-anchor="middle">digital_top_scoreboard</text><text class="small" x="775" y="299" text-anchor="middle">4 analysis exports</text><path class="line" d="M265 291H330"/><path class="line" d="M575 291H650"/>
<rect class="blue" x="65" y="380" width="200" height="80" rx="6"/><text x="165" y="410" text-anchor="middle">Chopper tap</text><text class="small" x="165" y="435" text-anchor="middle">chopper_intf_0</text>
<rect class="green" x="330" y="380" width="245" height="80" rx="6"/><text x="452" y="410" text-anchor="middle">chopper_monitor</text><text class="small" x="452" y="435" text-anchor="middle">chopper_dump_aport_d</text>
<rect class="amber" x="650" y="380" width="250" height="80" rx="6"/><text x="775" y="410" text-anchor="middle">chopper_scoreboard</text><text class="small" x="775" y="435" text-anchor="middle">out_data_chop_aexport_d</text><path class="line" d="M265 420H330"/><path class="line" d="M575 420H650"/>
<rect class="blue" x="65" y="505" width="200" height="95" rx="6"/><text x="165" y="535" text-anchor="middle">Analog DUT taps</text><text class="small" x="165" y="559" text-anchor="middle">data / POL / unlock / VBK</text>
<rect class="green" x="330" y="505" width="245" height="95" rx="6"/><text x="452" y="535" text-anchor="middle">analog output monitor</text><text class="small" x="452" y="559" text-anchor="middle">4 analysis ports</text>
<rect class="amber" x="650" y="505" width="250" height="95" rx="6"/><text x="775" y="535" text-anchor="middle">analog output scoreboard</text><text class="small" x="775" y="559" text-anchor="middle">monitor/pol/unlock/vbk imp</text><path class="line" d="M265 552H330"/><path class="line" d="M575 552H650"/>
<rect class="box" x="950" y="125" width="180" height="475" rx="6"/><text x="1040" y="157" text-anchor="middle">Checker gates</text><text class="small" x="1040" y="186" text-anchor="middle">components exist</text><text class="small" x="1040" y="208" text-anchor="middle">runtime cfg controls</text><text class="small" x="1040" y="230" text-anchor="middle">comparison behavior</text><text class="small" x="1040" y="575" text-anchor="middle">not build gating</text>
<text class="small" x="36" y="690">Source: checker_env/checker_agent.sv and monitor/scoreboard analysis-port declarations</text>
"""
    write_text(ASSET_DIR / "hk1v11-architecture.svg", svg_document(1200, 710, architecture, "HK1V11 UVM verification hierarchy"))
    write_text(ASSET_DIR / "hk1v11-checker-architecture.svg", svg_document(1200, 710, checkers, "HK1V11 checker topology"))


def parse_macros(text: str) -> dict[str, str]:
    """Parse active SystemVerilog preprocessor definitions."""
    macros: dict[str, str] = {}
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped.startswith("`define "):
            continue
        match = re.match(r"`define\s+(\w+)(?:\s+(.*?))?\s*$", stripped)
        if match:
            macros[match.group(1)] = (match.group(2) or "1").strip()
    return macros


def parse_assignments(text: str) -> dict[str, str]:
    """Parse direct environment configuration assignments."""
    assignments: dict[str, str] = {}
    pattern = re.compile(r"(?:rx_cfg|env_cfg)\.(\w+)\s*=\s*([^;\n]+)")
    for key, value in pattern.findall(text):
        assignments[key] = " ".join(value.split())
    return assignments


def classify_source_case(name: str, source: str) -> str:
    """Classify a source testcase using stable naming features."""
    key = name.lower()
    if "checksum" in key:
        return "Checksum"
    if "i2c" in key or "reg_" in key:
        return "Register / I2C"
    if "dplc" in key:
        return "DPLC"
    if any(token in key for token in ("chop", "vgma", "dbc", "smart_tp")):
        return "Analog / Chopper"
    if any(token in key for token in ("error", "prefix", "training", "bac_bac")):
        return "Protocol abnormal"
    if any(token in key for token in ("wake", "unlock", "rst", "power")):
        return "Power / Reset"
    if "force " in source:
        return "Directed waveform"
    return "Datapath"


def scan_source_cases() -> list[dict[str, object]]:
    """Scan testcase directories and collect source-derived metadata."""
    active_names = set()
    case_list = TESTS_DIR / "case_list.txt"
    if case_list.exists():
        active_names = {
            line.strip()
            for line in read_text(case_list).splitlines()
            if line.strip() and not line.lstrip().startswith("#")
        }
    result: list[dict[str, object]] = []
    if not TESTS_DIR.exists():
        return result
    for case_dir in sorted(TESTS_DIR.glob("t_*"), key=lambda item: item.name.lower()):
        if not case_dir.is_dir():
            continue
        preferred = case_dir / f"{case_dir.name}.sv"
        candidates = [
            path
            for path in case_dir.glob("*.sv")
            if path.name not in {"user_def.sv", "test_lib.sv", "waves_dumper.sv"}
        ]
        source_path = preferred if preferred.exists() else (
            max(candidates, key=lambda path: path.stat().st_size) if candidates else None
        )
        source = read_text(source_path) if source_path else ""
        macros = parse_macros(read_text(case_dir / "user_def.sv"))
        assignments = parse_assignments(source)
        checks = sorted(
            key.removesuffix("_check_on")
            for key, value in assignments.items()
            if key.endswith("_check_on") and value in {"1", "1'b1"}
        )
        disabled_checks = sorted(
            key.removesuffix("_check_on")
            for key, value in assignments.items()
            if key.endswith("_check_on") and value in {"0", "1'b0"}
        )
        result.append(
            {
                "name": case_dir.name,
                "category": classify_source_case(case_dir.name, source),
                "active": case_dir.name in active_names,
                "source": (
                    str(source_path.relative_to(SOURCE_ROOT)).replace("\\", "/")
                    if source_path
                    else "No testcase .sv"
                ),
                "macros": {
                    key: macros[key]
                    for key in (
                        "CHIP_SEL", "COLOR_DEPTH", "PORT_NUM", "PAIR_NUM",
                        "HACT", "VACT", "REF_RATE", "FRAME_NUM",
                    )
                    if key in macros
                },
                "checks": checks,
                "disabledChecks": disabled_checks,
                "cfgCount": len(list(case_dir.glob("cfg_frame*.txt"))),
                "patternCount": len(list(case_dir.glob("pattern/*.ppm"))),
                "forceCount": len(re.findall(r"(?m)^\s*force\s+", source)),
                "checkCalls": len(re.findall(r"\b(?:check_|compare_)\w*\s*\(", source)),
                "errorCalls": len(re.findall(r"`uvm_(?:error|fatal)\b", source)),
            }
        )
    return result


def find_header(rows: list[list[str]], required: tuple[str, ...]) -> int:
    """Return the first row containing any required header token."""
    for index, row in enumerate(rows):
        joined = " ".join(row).lower()
        if any(token.lower() in joined for token in required):
            return index
    return 0


def sheet_rows(workbook, sheet_name: str) -> list[list[str]]:
    """Return normalized values for one workbook sheet."""
    sheet = workbook[sheet_name]
    return [[clean_text(value) for value in row] for row in sheet.iter_rows(values_only=True)]


def parse_case_sheet(rows: list[list[str]], sheet_name: str) -> list[dict[str, str]]:
    """Parse normal or waveform case sheets."""
    if not rows:
        return []
    header_index = find_header(rows, ("Case Name",))
    headers = [value.lower() for value in rows[header_index]]

    def column(*tokens: str, default: int = -1) -> int:
        for index, value in enumerate(headers):
            if any(token.lower() in value for token in tokens):
                return index
        return default

    name_col = column("case name", default=0)
    desc_col = column("description", default=1)
    check_col = column("check", default=2)
    owner_col = column("owner", default=4)
    comment_col = column("comment", default=-1)
    feature1_col = next(
        (index for index, value in enumerate(headers) if value.strip() == "feature i"),
        -1,
    )
    feature2_col = next(
        (index for index, value in enumerate(headers) if value.strip() == "feature ii"),
        -1,
    )
    status_cols = [index for index, value in enumerate(headers) if "status" in value]
    date_cols = [index for index, value in enumerate(headers) if value.strip() == "date"]
    result: list[dict[str, str]] = []
    current_feature1 = ""
    current_feature2 = ""
    for row in rows[header_index + 1 :]:
        if feature1_col >= 0 and row[feature1_col]:
            current_feature1 = row[feature1_col]
        if feature2_col >= 0 and row[feature2_col]:
            current_feature2 = row[feature2_col]
        name = row[name_col].strip() if name_col < len(row) else ""
        if not re.match(r"^t_[A-Za-z0-9_\[\]-]+$", name):
            continue
        statuses = [row[index] for index in status_cols if index < len(row) and row[index]]
        result.append(
            {
                "name": name,
                "sheet": sheet_name,
                "feature1": (
                    current_feature1
                    if feature1_col >= 0 and current_feature1
                    else ("Normal" if "normal" in sheet_name else "Waveform")
                ),
                "feature2": current_feature2 if feature2_col >= 0 else "",
                "feature3": "",
                "priority": "",
                "description": row[desc_col] if desc_col < len(row) else "",
                "check": row[check_col] if check_col < len(row) else "",
                "comment": row[comment_col] if 0 <= comment_col < len(row) else "",
                "owner": row[owner_col] if owner_col < len(row) else "",
                "status": " / ".join(statuses),
                "date": " / ".join(
                    row[index] for index in date_cols if index < len(row) and row[index]
                ),
            }
        )
    return result


def parse_format_sheet(rows: list[list[str]]) -> list[dict[str, str]]:
    """Parse the hierarchical case plan sheet."""
    if not rows:
        return []
    header_index = find_header(rows, ("Feature I",))
    result: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        padded = row + [""] * max(0, 16 - len(row))
        name = padded[5].strip()
        if not name.startswith("t_"):
            continue
        result.append(
            {
                "name": name,
                "sheet": "caselist format",
                "feature1": padded[0],
                "feature2": padded[1],
                "feature3": padded[2],
                "priority": padded[3],
                "description": padded[6],
                "check": padded[8],
                "owner": padded[10],
                "status": " / ".join(value for value in padded[12:16] if value),
            }
        )
    return result


def merge_plan_cases(records: list[dict[str, str]]) -> list[dict[str, str]]:
    """Merge duplicate plan rows without losing sheet-specific details."""
    merged: dict[str, dict[str, str]] = {}
    for record in records:
        key = record["name"].strip().lower()
        if key not in merged:
            merged[key] = record.copy()
            continue
        current = merged[key]
        for field in (
            "feature1", "feature2", "feature3", "priority", "description",
            "check", "comment", "owner", "status", "date",
        ):
            if not current.get(field) and record.get(field):
                current[field] = record[field]
        sheets = {item.strip() for item in current["sheet"].split("+")}
        sheets.add(record["sheet"])
        current["sheet"] = " + ".join(sorted(sheets))
    return sorted(merged.values(), key=lambda item: item["name"].lower())


def parse_simple_table(rows: list[list[str]]) -> dict[str, object]:
    """Return a generic non-empty table representation."""
    non_empty = [row for row in rows if any(row)]
    if not non_empty:
        return {"headers": [], "rows": []}
    width = max(len(row) for row in non_empty)
    headers = non_empty[0] + [""] * (width - len(non_empty[0]))
    body = [row + [""] * (width - len(row)) for row in non_empty[1:]]
    return {"headers": headers, "rows": body}


def load_plan() -> dict[str, object]:
    """Load testcase, status, coverage, and environment data from Excel."""
    workbook = load_workbook(CASELIST_XLSX, read_only=True, data_only=True)
    sheets = {name: sheet_rows(workbook, name) for name in workbook.sheetnames}
    case_records: list[dict[str, str]] = []
    for name in ("normal case", "waveform case"):
        if name in sheets:
            case_records.extend(parse_case_sheet(sheets[name], name))
    return {
        "cases": merge_plan_cases(case_records),
        "environment": parse_simple_table(sheets.get("ENV changelist", [])),
        "status": parse_simple_table(sheets.get("case status ", [])),
        "coverage": parse_simple_table(sheets.get("coverage", [])),
        "videoFormats": parse_simple_table(sheets.get("video_format", [])),
        "registerRows": max(0, len(sheets.get("register", [])) - 1),
        "history": parse_simple_table(sheets.get("History", [])),
        "sheetNames": workbook.sheetnames,
    }


def link_cases(plan: dict[str, object], source_cases: list[dict[str, object]]) -> None:
    """Annotate plan and source cases using normalized exact-name matching."""
    source_map = {str(case["name"]).strip().lower(): case for case in source_cases}
    plan_map = {
        str(case["name"]).strip().lower(): case
        for case in plan["cases"]  # type: ignore[index]
    }
    for case in plan["cases"]:  # type: ignore[index]
        source = source_map.get(str(case["name"]).strip().lower())
        case["hasSource"] = source is not None
        case["source"] = source["source"] if source else ""
        case["sourceCategory"] = source["category"] if source else ""
        case["sourceMeta"] = source or {}
    for source in source_cases:
        source["inPlan"] = str(source["name"]).strip().lower() in plan_map


def render_table(table: dict[str, object], limit: int | None = None) -> str:
    """Render a generic table with empty columns removed."""
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    if not headers or not rows:
        return '<p class="empty">无可用数据。</p>'
    visible = [
        index
        for index in range(len(headers))
        if headers[index] or any(index < len(row) and row[index] for row in rows)
    ]
    display_rows = rows[:limit] if limit else rows
    head = "".join(f"<th>{esc(headers[index] or f'列 {index + 1}')}</th>" for index in visible)
    body = "".join(
        "<tr>" + "".join(
            f"<td>{esc(row[index]).replace(chr(10), '<br>')}</td>" for index in visible
        ) + "</tr>"
        for row in display_rows
    )
    return f'<div class="table-wrap"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>'


def prepare_page_toc(body: str) -> tuple[str, str]:
    """Add stable heading IDs and build a compact page table of contents."""
    heading_index = 0

    def add_id(match: re.Match[str]) -> str:
        nonlocal heading_index
        tag = match.group(1)
        attributes = match.group(2) or ""
        content = match.group(3)
        if re.search(r"\bid=", attributes):
            return match.group(0)
        heading_index += 1
        return f'<{tag}{attributes} id="section-{heading_index}">{content}</{tag}>'

    body = re.sub(
        r"<(h[23])(\s[^>]*)?>([\s\S]*?)</\1>",
        add_id,
        body,
        flags=re.IGNORECASE,
    )
    links: list[str] = []
    seen: set[str] = set()
    for match in re.finditer(
        r'<(h[23])[^>]*\bid="([^"]+)"[^>]*>([\s\S]*?)</\1>',
        body,
        flags=re.IGNORECASE,
    ):
        heading, section_id, content = match.groups()
        if section_id in seen:
            continue
        seen.add(section_id)
        label = re.sub(r"<[^>]+>", "", content).strip()
        if label:
            links.append(
                f'<a class="toc-{heading.lower()}" href="#{esc(section_id)}">{esc(label)}</a>'
            )
    return body, "".join(links[:30])


def page_shell(title: str, active: str, body: str, lead: str) -> str:
    """Wrap page content in the shared guide layout."""
    body, toc_links = prepare_page_toc(body)
    nav = "".join(
        f'<a href="{href}" class="{"active" if href == active else ""}">{label}</a>'
        for href, label in NAV_ITEMS
    )
    return f"""<!DOCTYPE html>
<!-- Auto-generated by tools/build_guide.py. -->
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{esc(title)} - HK1V11 DV 指南</title>
  <link rel="stylesheet" href="assets/guide.css">
</head>
<body>
  <header class="topbar">
    <a class="brand" href="index.html">HK1V11 DV 指南</a>
    <nav>{nav}</nav>
  </header>
  <main>
    <section class="page-head">
      <p class="eyebrow">HK1V11 · Verification Reference</p>
      <h1>{esc(title)}</h1>
      <p>{esc(lead)}</p>
    </section>
    <div class="content-grid">
      <aside class="section-nav"><strong>本页目录</strong>{toc_links}</aside>
      <div class="page-content">{body}</div>
    </div>
  </main>
  <footer>数据构建日期：{BUILD_DATE} · 来源：HK1V11 EDA case list 与 DV_TCON_C</footer>
  <script src="assets/guide.js"></script>
</body>
</html>
"""


def build_cases_page(plan: dict[str, object], source_cases: list[dict[str, object]]) -> str:
    """Build the searchable testcase plan and source index page."""
    plan_cases = plan["cases"]  # type: ignore[index]
    linked = sum(bool(case["hasSource"]) for case in plan_cases)
    active = sum(bool(case["active"]) for case in source_cases)
    categories = sorted({str(case["feature2"] or case["sourceCategory"] or "未分类") for case in plan_cases})
    legacy_asset = ASSET_DIR / "legacy" / "cases-01.svg"
    legacy = ""
    if legacy_asset.exists():
        legacy = """
<section class="band" id="lifecycle">
  <div class="section-head"><div><span>Legacy asset</span><h2>原有 testcase 创建流程图</h2></div></div>
  <figure class="diagram"><img src="assets/legacy/cases-01.svg" alt="原有 testcase 创建流程图"></figure>
</section>"""
    category_options = "".join(f'<option value="{esc(value)}">{esc(value)}</option>' for value in categories)
    body = f"""
<section class="metrics">
  <div><strong>{len(plan_cases)}</strong><span>Excel 唯一计划 case</span></div>
  <div><strong>{len(source_cases)}</strong><span>Excel case 已匹配源码</span></div>
  <div><strong>{linked}</strong><span>计划与源码精确关联</span></div>
  <div><strong>{active}</strong><span>匹配且 case_list 启用</span></div>
</section>
<section class="band" id="index">
  <div class="section-head"><div><span>Case browser</span><h2>验证计划 testcase</h2></div><p>名称采用去空格、忽略大小写的精确匹配，不使用模糊包含关系。</p></div>
  <div class="filters" data-case-filters>
    <label>搜索<input type="search" id="case-search" placeholder="Case、Feature、描述、Owner"></label>
    <label>分类<select id="case-category"><option value="">全部分类</option>{category_options}</select></label>
    <label>源码<select id="case-source"><option value="">全部</option><option value="yes">已有源码</option><option value="no">缺少源码</option></select></label>
    <label>来源<select id="case-sheet"><option value="">全部 sheet</option><option value="normal case">normal case</option><option value="waveform case">waveform case</option><option value="caselist format">caselist format</option></select></label>
  </div>
  <p class="result-line"><strong id="case-visible">{len(plan_cases)}</strong> / {len(plan_cases)} 条计划记录</p>
  <div id="case-list" class="case-list"></div>
</section>
<section class="band" id="source-only">
  <div class="section-head"><div><span>Source audit</span><h2>Excel case 对应源码</h2></div><p>仅列出 Excel 主清单中存在且已匹配源码的 testcase，不展示计划外目录。</p></div>
  <div id="source-case-list" class="source-grid"></div>
</section>
{legacy}
<script src="assets/cases-data.js"></script>
"""
    return page_shell(
        "Testcase 计划与源码索引",
        "cases.html",
        body,
        "以 3.HK1V11_EDA_case_list.xlsx 为计划主清单，并与当前 DV_TCON_C testcase 源码逐项核对。",
    )


def build_plan_page(plan: dict[str, object], source_cases: list[dict[str, object]]) -> str:
    """Build the verification plan dashboard."""
    cases = plan["cases"]  # type: ignore[index]
    features1 = Counter(case["feature1"] or "未分类" for case in cases)
    features2 = Counter(case["feature2"] or "未分类" for case in cases)
    linked = sum(bool(case["hasSource"]) for case in cases)
    feature_rows = "".join(
        f"<tr><td>{esc(name)}</td><td>{count}</td></tr>"
        for name, count in features1.most_common()
    )
    feature2_rows = "".join(
        f"<tr><td>{esc(name)}</td><td>{count}</td></tr>"
        for name, count in features2.most_common()
    )
    body = f"""
<section class="metrics">
  <div><strong>{len(cases)}</strong><span>唯一计划 case</span></div>
  <div><strong>{linked}</strong><span>已关联源码</span></div>
  <div><strong>{len(cases) - linked}</strong><span>缺少同名源码</span></div>
  <div><strong>{plan['registerRows']}</strong><span>register sheet 数据行</span></div>
</section>
<section class="band" id="coverage-map">
  <div class="section-head"><div><span>Plan hierarchy</span><h2>Feature 覆盖结构</h2></div><p>统计来自 Excel 计划行，不将 testcase 目录数量当成计划覆盖率。</p></div>
  <div class="two-col">
    <div><h3>Feature I / Case 类型</h3><table><thead><tr><th>分类</th><th>Case 数</th></tr></thead><tbody>{feature_rows}</tbody></table></div>
    <div><h3>Feature II / Spec 模块</h3><table><thead><tr><th>模块</th><th>Case 数</th></tr></thead><tbody>{feature2_rows}</tbody></table></div>
  </div>
</section>
<section class="band" id="status">
  <div class="section-head"><div><span>Execution state</span><h2>Case 状态汇总</h2></div></div>
  {render_table(plan['status'])}
</section>
<section class="band" id="coverage">
  <div class="section-head"><div><span>Coverage snapshot</span><h2>覆盖率快照</h2></div><p>保留 Excel 中的原始统计口径与日期。</p></div>
  {render_table(plan['coverage'])}
</section>
<section class="band" id="environment">
  <div class="section-head"><div><span>Environment backlog</span><h2>环境修改清单</h2></div><p>状态和负责人直接取自 ENV changelist。</p></div>
  {render_table(plan['environment'])}
</section>
<section class="band" id="formats">
  <div class="section-head"><div><span>Product matrix</span><h2>Video format 与应用场景</h2></div><p>宽表支持横向滚动，保留应用、分辨率、带宽、pair 和 driver 信息。</p></div>
  {render_table(plan['videoFormats'])}
</section>
<section class="band" id="history">
  <div class="section-head"><div><span>Workbook governance</span><h2>计划变更历史</h2></div></div>
  {render_table(plan['history'])}
</section>
"""
    return page_shell(
        "验证计划",
        "plan.html",
        body,
        "集中展示 Excel case 主清单、环境修改、执行状态、覆盖率及产品 video format。",
    )


def scan_scripts() -> list[dict[str, str]]:
    """Scan current automation scripts and infer their operational role."""
    patterns = ("*.py", "*.sh", "*.pl")
    result: list[dict[str, str]] = []
    for pattern in patterns:
        for path in SOURCE_ROOT.rglob(pattern):
            if any(part in {"simv.daidir", "csrc", ".git"} for part in path.parts):
                continue
            rel = str(path.relative_to(SOURCE_ROOT)).replace("\\", "/")
            key = path.name.lower()
            if "create" in key or "copy" in key:
                role = "Testcase 创建与复制"
            elif "reg" in key or "cfg" in key:
                role = "寄存器与配置生成"
            elif "dplc" in key or "ppm" in key or "pattern" in key:
                role = "Pattern / CModel"
            elif "cov" in key or "regression" in key or "multi" in key:
                role = "回归与 coverage"
            else:
                role = "仿真辅助"
            result.append({"path": rel, "type": path.suffix.lstrip(".").upper(), "role": role})
    return sorted(result, key=lambda item: item["path"].lower())


def build_scripts_page(scripts: list[dict[str, str]]) -> str:
    """Build the script inventory page."""
    rows = "".join(
        f"<tr><td><code>{esc(item['path'])}</code></td><td>{esc(item['type'])}</td><td>{esc(item['role'])}</td></tr>"
        for item in scripts
    )
    counts = Counter(item["type"] for item in scripts)
    body = f"""
<section class="metrics">
  <div><strong>{len(scripts)}</strong><span>脚本总数</span></div>
  <div><strong>{counts.get('PY', 0)}</strong><span>Python</span></div>
  <div><strong>{counts.get('SH', 0)}</strong><span>Shell</span></div>
  <div><strong>{counts.get('PL', 0)}</strong><span>Perl</span></div>
</section>
<section class="band"><div class="section-head"><div><span>Source inventory</span><h2>自动化脚本清单</h2></div><p>路径来自当前 HK1V11 DV_TCON_C，备份与仿真生成目录不计入。</p></div>
<div class="table-wrap"><table><thead><tr><th>路径</th><th>类型</th><th>用途</th></tr></thead><tbody>{rows}</tbody></table></div></section>
"""
    return page_shell("脚本指南", "scripts.html", body, "按实际源码路径整理 testcase、寄存器、pattern、回归及仿真辅助脚本。")


def build_css() -> str:
    """Return the shared visual system."""
    base = """/* Auto-generated by tools/build_guide.py. */
:root{--bg:#f4f6f8;--surface:#fff;--ink:#1b232b;--fg:#1b232b;--text:#1b232b;--muted:#66727d;--line:#d9dee3;--border:#d9dee3;--accent:#087e6b;--accent-soft:#dff3ee;--blue:#245c91;--pass:#16805f;--pass-soft:#e0f3eb;--warn:#a85d00;--warn-soft:#fff0d8;--code-bg:#17212b;--code-fg:#edf3f6;--radius:6px}*{box-sizing:border-box}html{scroll-behavior:smooth}body{margin:0;background:var(--bg);color:var(--ink);font-family:"Segoe UI","Microsoft YaHei",sans-serif;font-size:15px;line-height:1.65;letter-spacing:0}.topbar{position:sticky;top:0;z-index:20;display:flex;align-items:center;gap:24px;min-height:58px;padding:0 28px;background:#fff;border-bottom:1px solid var(--line)}.brand{font-weight:750;color:var(--ink);text-decoration:none;white-space:nowrap}.topbar nav{display:flex;gap:2px;overflow-x:auto}.topbar nav a{padding:17px 10px 14px;color:var(--muted);text-decoration:none;white-space:nowrap;border-bottom:3px solid transparent}.topbar nav a:hover,.topbar nav a.active{color:var(--accent);border-color:var(--accent)}main{max-width:1440px;margin:auto}.page-head{padding:42px 32px 30px;border-bottom:1px solid var(--line);background:#fff}.page-head h1{margin:4px 0 8px;font-size:32px;line-height:1.2}.page-head>p:last-child{max-width:900px;margin:0;color:var(--muted)}.eyebrow,.section-head span{text-transform:uppercase;color:var(--accent);font-size:12px;font-weight:750}.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));border-bottom:1px solid var(--line);background:#fff}.metrics div{padding:24px 30px;border-right:1px solid var(--line)}.metrics div:last-child{border-right:0}.metrics strong{display:block;font-size:30px;line-height:1.1}.metrics span{color:var(--muted);font-size:13px}.band{padding:34px 32px;border-bottom:1px solid var(--line)}.band:nth-of-type(even){background:#fff}.section-head{display:flex;justify-content:space-between;gap:28px;align-items:end;margin-bottom:20px}.section-head h2{margin:2px 0 0;font-size:23px}.section-head p{max-width:540px;margin:0;color:var(--muted);text-align:right}.two-col{display:grid;grid-template-columns:1fr 1fr;gap:24px}.two-col>div{min-width:0}.two-col h3{font-size:16px}.table-wrap{overflow:auto;max-height:680px;border:1px solid var(--line);background:#fff}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:9px 11px;text-align:left;vertical-align:top;border-bottom:1px solid var(--line);border-right:1px solid var(--line);min-width:90px}th{position:sticky;top:0;z-index:2;background:#edf2f4;color:#30404d}td{white-space:normal;overflow-wrap:anywhere}code{font-family:Consolas,monospace;font-size:12px;color:var(--blue)}.filters{display:grid;grid-template-columns:2fr repeat(3,1fr);gap:12px;padding:14px;background:#fff;border:1px solid var(--line);border-radius:var(--radius)}label{font-size:12px;color:var(--muted)}input,select{display:block;width:100%;height:38px;margin-top:4px;padding:0 10px;border:1px solid #bcc5cc;border-radius:4px;background:#fff;color:var(--ink)}.result-line{color:var(--muted)}.case-list{display:grid;gap:10px}.case-row{display:grid;grid-template-columns:minmax(250px,1.1fr) minmax(180px,.8fr) minmax(260px,2fr) 110px;gap:16px;padding:15px 16px;background:#fff;border:1px solid var(--line);border-left:4px solid var(--accent);border-radius:var(--radius)}.case-row.missing{border-left-color:var(--warn)}.case-row h3{margin:0 0 5px;font-size:14px;overflow-wrap:anywhere}.case-row p{margin:0;color:var(--muted);font-size:12px;white-space:pre-line}.case-row .description{color:var(--ink)}.pill{display:inline-block;padding:2px 7px;margin:2px 3px 2px 0;border-radius:999px;background:var(--accent-soft);color:var(--accent);font-size:11px}.pill.warn{background:var(--warn-soft);color:var(--warn)}.source-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(310px,1fr));gap:10px}.source-card{padding:14px;background:#fff;border:1px solid var(--line);border-radius:var(--radius)}.source-card h3{margin:0;font-size:13px;overflow-wrap:anywhere}.source-card p{margin:7px 0 0;color:var(--muted);font-size:12px}.diagram{margin:0;border:1px solid var(--line);background:#fff;overflow:auto}.diagram img{display:block;max-width:100%;height:auto;margin:auto}.legacy-content{padding:34px 32px}.legacy-content .hero{text-align:center;padding:12px 0 30px}.legacy-content .hero h1{font-size:28px;margin:0 0 8px}.legacy-content .author,.legacy-content .lead{color:var(--muted)}.legacy-content .badges{display:flex;justify-content:center;gap:8px;flex-wrap:wrap;margin-top:14px}.legacy-content .section{margin-bottom:38px}.legacy-content .section h2{font-size:23px;margin:0 0 18px}.legacy-content .section h3{font-size:17px;margin:24px 0 12px;border-left:4px solid var(--accent);padding-left:10px}.legacy-content .card{padding:18px;margin:12px 0;background:#fff;border:1px solid var(--line);border-radius:var(--radius)}.legacy-content .tag{display:inline-block;padding:3px 8px;border-radius:999px;background:var(--accent-soft);color:var(--accent);font-size:12px}.legacy-content .hint-bar,.legacy-content .blocker-list li{padding:13px 16px;background:var(--warn-soft);border-left:4px solid var(--warn);border-radius:var(--radius)}.legacy-content .tip-list li{padding:13px 16px;background:var(--pass-soft);border-left:4px solid var(--pass);border-radius:var(--radius)}.legacy-content .chapter-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:12px}.legacy-content .chapter-card,.legacy-content .page-nav a{display:block;padding:16px;background:#fff;border:1px solid var(--line);border-radius:var(--radius);color:var(--ink);text-decoration:none}.legacy-content .chapter-card .num{color:var(--accent);font-weight:700}.legacy-content .chapter-card .title{font-weight:700}.legacy-content .chapter-card .desc{color:var(--muted);font-size:13px}.legacy-content .page-nav{display:grid;grid-template-columns:1fr 1fr;gap:12px}.legacy-content pre{padding:16px;overflow:auto;background:var(--code-bg);color:var(--code-fg);border-radius:var(--radius)}.legacy-content .diagram{padding:18px;margin:12px 0}.legacy-content .diagram svg{display:block;max-width:100%;height:auto;margin:auto}.legacy-content table{background:#fff}footer{padding:28px;text-align:center;color:var(--muted);font-size:12px}.empty{color:var(--muted)}@media(max-width:900px){.topbar{padding:0 14px}.brand{display:none}.page-head{padding:28px 18px}.page-head h1{font-size:25px}.metrics{grid-template-columns:1fr 1fr}.metrics div:nth-child(2){border-right:0}.band,.legacy-content{padding:26px 18px}.section-head{display:block}.section-head p{text-align:left;margin-top:8px}.two-col{grid-template-columns:1fr}.filters{grid-template-columns:1fr 1fr}.case-row{grid-template-columns:1fr}.case-row>div{min-width:0}.legacy-content .page-nav{grid-template-columns:1fr}}@media(max-width:520px){.metrics{grid-template-columns:1fr}.metrics div{border-right:0}.filters{grid-template-columns:1fr}.topbar nav a{padding-left:8px;padding-right:8px}.page-head h1{font-size:23px}}
"""
    return base + """
.content-grid{display:grid;grid-template-columns:48px minmax(0,1fr);align-items:start}.page-content{min-width:0}.section-nav{position:sticky;top:70px;z-index:12;width:44px;max-height:calc(100vh - 86px);overflow:hidden;padding:8px 4px;background:var(--surface);border:1px solid var(--line);border-radius:6px;transition:width .18s ease,box-shadow .18s ease}.section-nav:hover,.section-nav:focus-within{width:230px;overflow-y:auto;box-shadow:0 8px 28px rgba(20,36,48,.14)}.section-nav strong{display:block;width:34px;color:var(--blue);font-size:13px;line-height:1.15;text-align:center}.section-nav:hover strong,.section-nav:focus-within strong{width:auto;text-align:left;margin:0 7px 8px}.section-nav a{display:block;padding:5px 7px;color:var(--muted);text-decoration:none;font-size:12px;line-height:1.35;border-radius:4px;white-space:normal}.section-nav:not(:hover):not(:focus-within) a{opacity:0;pointer-events:none}.section-nav a:hover{color:var(--blue);background:#e8f0f7}.section-nav .toc-h3{padding-left:16px}.page-content h2[id],.page-content h3[id]{scroll-margin-top:76px}.case-list{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}.case-card{min-width:0;padding:18px;background:var(--surface);border:1px solid var(--line);border-radius:6px}.case-card.missing{border-left:4px solid var(--warn)}.case-title{display:flex;align-items:flex-start;justify-content:space-between;gap:10px}.case-title code{font-size:13px;overflow-wrap:anywhere}.case-tags{display:flex;gap:5px;flex-wrap:wrap;margin:9px 0}.case-card h4{margin:14px 0 4px;font-size:13px}.case-description{margin:3px 0;color:var(--muted);font-size:13px;white-space:pre-line}.case-checkpoint{margin:4px 0 10px;padding:10px 12px;border-left:4px solid var(--accent);background:var(--accent-soft);font-size:13px;white-space:pre-line}.pill.pass{border:1px solid var(--pass);background:var(--pass-soft);color:var(--pass)}.case-evidence{margin-top:9px;font-size:12px}.case-evidence summary{cursor:pointer;color:var(--blue);font-weight:650}.case-evidence pre{max-height:280px;margin:8px 0}.source-upgrade{padding:32px;background:#eef3f5;border-bottom:1px solid var(--line)}.source-upgrade h2{margin-top:0}.source-upgrade .diagram img{width:100%;min-width:760px}@media(max-width:900px){.content-grid{grid-template-columns:1fr}.section-nav{position:static;width:auto;max-height:150px;margin:12px 14px;padding:10px;overflow-y:auto}.section-nav:hover,.section-nav:focus-within{width:auto}.section-nav strong,.section-nav:hover strong{width:auto;text-align:left;margin:0 7px 8px}.section-nav:not(:hover):not(:focus-within) a{display:inline-block;opacity:1;pointer-events:auto}.case-list{grid-template-columns:1fr}.source-upgrade{padding:24px 18px}}
"""


def build_js() -> str:
    """Return the shared case browser behavior."""
    return """// Auto-generated by tools/build_guide.py.
(() => {
  const data = window.HK1V11_CASE_DATA;
  if (!data) return;
  const list = document.querySelector('#case-list');
  const sourceList = document.querySelector('#source-case-list');
  const visible = document.querySelector('#case-visible');
  const escapeHtml = (value) => String(value ?? '').replace(/[&<>\"']/g, (char) => ({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}[char]));
  const tags = (values, prefix = '') => values.map((value) => `<span class="pill">${escapeHtml(prefix + value)}</span>`).join('');
  const renderCases = () => {
    const query = document.querySelector('#case-search').value.trim().toLowerCase();
    const category = document.querySelector('#case-category').value;
    const source = document.querySelector('#case-source').value;
    const sheet = document.querySelector('#case-sheet').value;
    const rows = data.planCases.filter((item) => {
      const haystack = [item.name,item.feature1,item.feature2,item.feature3,item.description,item.owner,item.status].join(' ').toLowerCase();
      const itemCategory = item.feature2 || item.sourceCategory || '未分类';
      return (!query || haystack.includes(query)) && (!category || itemCategory === category) && (!source || (source === 'yes') === item.hasSource) && (!sheet || item.sheet.includes(sheet));
    });
    visible.textContent = rows.length;
    if (!rows.length) {
      list.innerHTML = '<div class="empty">没有匹配的 testcase</div>';
      return;
    }
    list.innerHTML = rows.slice(0, 500).map((item) => {
      const meta = item.sourceMeta || {};
      const macroTags = Object.entries(meta.macros || {}).map(([key, value]) => `${key}=${value}`);
      const evidence = [
        `Excel sheet: ${item.sheet}`,
        `Source: ${item.source || 'not linked'}`,
        `cfg_frame files: ${meta.cfgCount || 0}`,
        `pattern files: ${meta.patternCount || 0}`,
        `force statements: ${meta.forceCount || 0}`,
        `check/compare calls: ${meta.checkCalls || 0}`,
        `uvm_error/fatal calls: ${meta.errorCalls || 0}`,
      ];
      const sourceDescription = item.hasSource
        ? `${item.sourceCategory || 'Source testcase'}；包含 ${meta.cfgCount || 0} 个 cfg_frame、${meta.patternCount || 0} 个 pattern，源码中有 ${meta.forceCount || 0} 处 force。`
        : 'Excel 中存在该计划项，但当前选定的 E 盘源码树中没有同名 testcase 目录。';
      return `<article class="case-card ${item.hasSource ? '' : 'missing'}">
        <div class="case-title"><code>${escapeHtml(item.name)}</code><span class="pill ${String(item.status).toUpperCase().includes('PASS') ? 'pass' : ''}">${escapeHtml(item.status || 'Status 未填写')}</span></div>
        <div class="case-tags"><span class="pill warn">Case Type: ${escapeHtml(item.sheet.includes('waveform') ? 'Waveform' : 'Normal')}</span><span class="pill">Feature I: ${escapeHtml(item.feature1 || 'Excel 未填写')}</span><span class="pill">Feature II: ${escapeHtml(item.feature2 || 'Excel 未填写')}</span><span class="pill ${item.hasSource ? 'pass' : 'warn'}">${item.hasSource ? 'source linked' : 'Excel only'}</span></div>
        <h4>验证目标</h4><p class="case-description">${escapeHtml(item.description || 'Excel 未填写验证目标')}</p>
        <h4>检查点</h4><p class="case-checkpoint">${escapeHtml(item.check || 'Excel 未填写检查点')}</p>
        ${item.comment ? `<h4>说明</h4><p class="case-description">${escapeHtml(item.comment)}</p>` : ''}
        <div class="case-tags">${item.owner ? `<span class="pill">Owner: ${escapeHtml(item.owner)}</span>` : ''}${item.date ? `<span class="pill">Date: ${escapeHtml(item.date)}</span>` : ''}<span class="pill">cfg: ${meta.cfgCount || 0}</span><span class="pill">pattern: ${meta.patternCount || 0}</span></div>
        <div class="case-tags">${tags(meta.checks || [], 'ON: ')}${tags(meta.disabledChecks || [], 'OFF: ')}${tags(macroTags)}</div>
        <details class="case-evidence"><summary>查看源码补充描述</summary><p>${escapeHtml(sourceDescription)}</p></details>
        <details class="case-evidence"><summary>查看 Excel 与源码证据</summary><pre><code>${escapeHtml(evidence.join('\\n'))}</code></pre></details>
      </article>`;
    }).join('');
  };
  document.querySelectorAll('[data-case-filters] input,[data-case-filters] select').forEach((control) => control.addEventListener('input', renderCases));
  renderCases();
  if (sourceList) sourceList.innerHTML = data.sourceCases.map((item) => `<article class="source-card"><h3>${escapeHtml(item.name)}</h3><p>${escapeHtml(item.category)} · ${item.active ? 'case_list 启用' : '未启用'}</p><p><code>${escapeHtml(item.source)}</code></p></article>`).join('');
})();
"""


def main() -> None:
    """Build generated guide pages and data assets."""
    if not CASELIST_XLSX.exists():
        raise FileNotFoundError(f"Case plan not found: {CASELIST_XLSX}")
    if not TESTS_DIR.exists():
        raise FileNotFoundError(f"Test source directory not found: {TESTS_DIR}")
    archive_legacy_svgs()
    archive_legacy_content()
    plan = load_plan()
    source_cases = scan_source_cases()
    link_cases(plan, source_cases)
    planned_source_cases = [case for case in source_cases if case["inPlan"]]
    scripts = scan_scripts()
    data = {"planCases": plan["cases"], "sourceCases": planned_source_cases}
    write_text(ASSET_DIR / "cases-data.js", "// Auto-generated by tools/build_guide.py.\nwindow.HK1V11_CASE_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    write_text(ASSET_DIR / "guide.css", build_css())
    write_text(ASSET_DIR / "guide.js", build_js())
    build_source_diagrams()
    build_legacy_pages()
    write_text(GUIDE_DIR / "cases.html", build_cases_page(plan, planned_source_cases))
    write_text(GUIDE_DIR / "plan.html", build_plan_page(plan, source_cases))
    write_text(GUIDE_DIR / "scripts.html", build_scripts_page(scripts))
    linked = sum(bool(case["hasSource"]) for case in plan["cases"])
    print(
        f"Built HK1V11 guide: {len(plan['cases'])} Excel plan cases, "
        f"{len(source_cases)} source directories scanned, {linked} Excel-listed "
        f"source cases exposed, {len(scripts)} scripts."
    )


if __name__ == "__main__":
    main()
